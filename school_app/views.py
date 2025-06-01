from django.shortcuts import render, redirect
from django.http import JsonResponse
import json
from .models import Class, Teacher, Teacher_Subject_Class_Relation,PeriodTiming,Student,Exam,SkillOrDevelopmentArea
from django.shortcuts import get_object_or_404
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

from django.shortcuts import render, redirect
from .models import MediaFile
from django.http import HttpResponse
from django.core.management import call_command

def run_migrations(request):
    try:
        call_command('migrate')
        return HttpResponse("Migrations applied successfully.")
    except Exception as e:
        return HttpResponse(f"Migration failed: {str(e)}")
from django.http import JsonResponse
from django.db import connection
from django.http import JsonResponse
from django.db import connection
from django.http import JsonResponse
from django.db import connection

def check_columns(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'school_app_student'")
        columns = [row[0] for row in cursor.fetchall()]
    return JsonResponse({'columns': columns})

def list_migrations(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT app, name, applied FROM django_migrations ORDER BY applied DESC;")
        migrations = cursor.fetchall()
    return JsonResponse({"applied_migrations": migrations})

from django.utils.timezone import now
from .models import MediaFile, SchoolProfile

def index_view(request):
    latest_media = MediaFile.objects.order_by('-uploaded_at')[:10]
    school_profile = SchoolProfile.objects.first()

    return render(request, 'school_management/index.html', {
        'media_files': latest_media,
        'year': now().year,
        'school_profile': school_profile,
    })


from django.shortcuts import render, redirect
from django.core.files.uploadedfile import InMemoryUploadedFile
from PIL import Image
import io

from django.shortcuts import render
from .models import MediaFile
from django.shortcuts import render, redirect
from django.core.files.uploadedfile import InMemoryUploadedFile
import io
from .models import MediaFile  # Make sure your model is correctly imported

def media_upload(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        media_type = request.POST.get('media_type')

        # Handle the cropped image (image file sent from frontend)
        cropped_image = request.FILES.get('image_file')
        if cropped_image:
            # Read the image data into memory
            image_data = cropped_image.read()

            # Create an in-memory file
            image_file = InMemoryUploadedFile(
                io.BytesIO(image_data),  # Use image data
                'ImageField',            # Field type
                cropped_image.name,      # File name
                'image/jpeg',            # Content type
                len(image_data),         # Size
                None                     # Charset
            )

            # Save the media file in the database
            MediaFile.objects.create(
                title=title,
                media_type=media_type,
                file=image_file
            )
            return redirect('media_gallery')  # Redirect to gallery

    return render(request, 'school_management/media_center/upload.html')

def media_gallery(request):
    media_files = MediaFile.objects.all()  # Fetch all media files from the database
    return render(request, 'school_management/media_center/gallery.html', {'media_files': media_files})
from django.shortcuts import get_object_or_404, redirect
from .models import MediaFile

def delete_media(request, media_id):
    media = get_object_or_404(MediaFile, id=media_id)
    media.delete()
    return redirect('media_gallery')

# Load Admin Dashboard
def admin_dashboard(request):
    return render(request, "school_management/excce'administration_panel.html")
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Class
def class_creation_pg(request):
    classes = Class.objects.all().order_by("name")  # Fetch classes sorted by name
    return render(request, "school_management/class_creation.html", {"classes": classes})
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Class

@csrf_exempt  # Only use for debugging. Use CSRF tokens in production.
def create_class(request):
    if request.method == "POST":
        try:
            # Debugging: Print the received request data
            print("Request Content-Type:", request.content_type)
            print("Request Body:", request.body.decode("utf-8"))

            if request.content_type == "application/json":
                data = json.loads(request.body)
            else:
                data = request.POST

            class_name = data.get("name", "").strip()
            if not class_name:
                return JsonResponse({"success": False, "message": "Class name cannot be empty!"}, status=400)

            new_class, created = Class.objects.get_or_create(name=class_name)
            if created:
                return JsonResponse({"success": True, "message": f"Class '{new_class.name}' created successfully!"})
            else:
                return JsonResponse({"success": False, "message": "Class already exists!"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON data"}, status=400)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=405)
def class_subject_management_pg(request):
    classes = Class.objects.all()
    subjects = Subject.objects.all()
    teacher_subject_class_relations = Teacher_Subject_Class_Relation.objects.all()
    return render(request, "school_management/class_subject_management.html", {
        "classes": classes,
        "subjects": subjects,
        "teacher_subject_class_relations": teacher_subject_class_relations
    })


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import get_object_or_404
from .models import Subject, Teacher_Subject_Class_Relation

@csrf_exempt  # Remove in production, use CSRF tokens
def delete_subject(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            subject_id = data.get("subject_id")

            if not subject_id:
                return JsonResponse({"success": False, "message": "Subject ID is required!"}, status=400)

            subject = get_object_or_404(Subject, id=subject_id)

            # Remove subject from Teacher_Subject_Class_Relation
            Teacher_Subject_Class_Relation.objects.filter(subject=subject).delete()
            # Delete the subject
            subject.delete()

            return JsonResponse({"success": True, "message": "Subject deleted successfully!"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=400)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=405)



from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
import json
import re
from .models import Student, Class

def student_enrollment(request):
    """Render the student enrollment form"""
    classes = Class.objects.all()
    return render(request, "school_management/student_enrollment.html", {"classes": classes})


import random
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Parent, Student, Class
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.hashers import make_password
import random
import string
import logging

from .models import Student, Parent, Class, AcademicYear

logger = logging.getLogger(__name__)


# Utility: Generate a unique parent ID (e.g., p1234)
def generate_parent_id():
    while True:
        parent_id = f"p{random.randint(1000, 9999)}"
        if not Parent.objects.filter(username=parent_id).exists():
            return parent_id


# Utility: Generate a simple password like "p1234p"
def generate_parent_password():
    digits = ''.join(random.choices(string.digits, k=4))
    return f"p{digits}p"


# Page loader for student enrollment
def student_enrollment(request):
    classes = Class.objects.all()
    academic_years = AcademicYear.objects.order_by('-start_date')
    return render(
        request,
        "school_management/student_enrollment.html",
        {"classes": classes, "academic_years": academic_years}
    )


# Student enrollment handler
def enroll_student(request):
    if request.method == "POST":
        try:
            # Extract form data
            name = request.POST.get("name", "").strip()
            parent_name = request.POST.get("parent_name", "").strip()
            phone = request.POST.get("phone", "").strip()
            aadhar = request.POST.get("aadhar", "").strip()
            email = request.POST.get("email", "").strip() or None
            address = request.POST.get("address", "").strip()
            category = request.POST.get("category", "General").strip()
            class_id = request.POST.get("assigned_class")
            academic_year_id = request.POST.get("academic_year")

            # Validate required fields
            if not all([name, parent_name, phone, class_id, aadhar, academic_year_id]):
                return JsonResponse({"error": "All required fields must be filled!"}, status=400)

            # Validate existence of related records
            assigned_class = get_object_or_404(Class, id=class_id)
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)

            # Generate login credentials for parent
            parent_password = generate_parent_password()
            parent_id = generate_parent_id()

            # Create or fetch parent
            parent, created = Parent.objects.get_or_create(
                phone=phone,
                defaults={
                    "name": parent_name,
                    "username": parent_id,
                    "aadhar": aadhar,
                    "email": email,
                    "address": address,
                    "password": make_password(parent_password)
                }
            )

            # Create student record
            student = Student.objects.create(
                name=name,
                phone=phone,
                aadhar=aadhar,
                email=email,
                address=address,
                assigned_class=assigned_class,
                academic_year=academic_year,
                category=category,
                parent=parent
            )

            return JsonResponse({
                "success": f"Student {name} enrolled successfully!",
                "roll_number": student.roll_number,
                "parent_id": parent.username,
                "password": parent_password if created else "Already registered"
            })

        except Exception as e:
            logger.error(f"Error enrolling student: {str(e)}", exc_info=True)
            return JsonResponse({"error": "Something went wrong. Please check input data or database setup."}, status=500)

    return JsonResponse({"error": "Invalid request method!"}, status=405)

from django.shortcuts import render
from .models import Class, Subject, Teacher_Subject_Class_Relation

from .models import Class, Subject, Teacher_Subject_Class_Relation, ClassSubSubjectAssignment
def teacher_enrollment_pg(request):
    subjects = Subject.objects.all()
    classes = Class.objects.all()
    assignments = Teacher_Subject_Class_Relation.objects.select_related("teacher", "subject").prefetch_related("assigned_classes")

    class_subject_teacher_data = []

    for class_obj in classes:
        row_data = {"class_name": class_obj.name, "subjects": []}

        for subject in subjects:
            # Check if any relation exists for this class and subject
            relations = assignments.filter(subject=subject, assigned_classes__id=class_obj.id)

            if relations.exists():
                teacher_names_list = [relation.teacher.name for relation in relations if relation.teacher]
                teacher_names = "<br>".join(teacher_names_list) if teacher_names_list else "Empty"
            else:
                teacher_names = "N/A"  # Mark as "N/A" when the subject is not assigned to the class

            row_data["subjects"].append({"subject_name": subject.name, "teacher_names": teacher_names})

        class_subject_teacher_data.append(row_data)

    return render(
        request, 
        "school_management/teacher_enrollment.html", 
        {
            "subjects": subjects, 
            "classes": classes, 
            "class_subject_teacher_data": class_subject_teacher_data,
            "assignments": assignments,
        }
    )


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Teacher_Subject_Class_Relation

@csrf_exempt
def delete_class_subj_Tr(request):
    if request.method == "POST":
        assignment_id = request.POST.get("assignment_id")
        try:
            assignment = Teacher_Subject_Class_Relation.objects.get(id=assignment_id)
            assignment.delete()
            return JsonResponse({"status": "success"})
        except Teacher_Subject_Class_Relation.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Assignment not found"})

    return JsonResponse({"status": "error", "message": "Invalid request"})


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Teacher, Subject, Class, Teacher_Subject_Class_Relation
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Teacher, Subject, Class, Teacher_Subject_Class_Relation
from django.db.utils import IntegrityError

@csrf_exempt  # Remove this in production
def enroll_teacher(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone_number = request.POST.get("phone_number")
        subject_ids = request.POST.getlist("subjects")
        class_ids = request.POST.getlist("assigned_classes")

        if not name or not phone_number or not subject_ids or not class_ids:
            return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)

        try:
            # ✅ Ensure the phone number is unique
            teacher = Teacher.objects.create(name=name, phone=phone_number)

            for subject_id in subject_ids:
                subject = Subject.objects.get(id=subject_id)

                relation = Teacher_Subject_Class_Relation.objects.create(
                    teacher=teacher,
                    subject=subject
                )
                relation.assigned_classes.set(Class.objects.filter(id__in=class_ids))
                relation.save()

            return JsonResponse({"status": "success", "message": "Teacher enrolled successfully"})

        except IntegrityError:
            return JsonResponse({"status": "error", "message": "Phone number already exists!"}, status=400)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

from django.http import JsonResponse
from .models import Class, Subject, Teacher_Subject_Class_Relation

def get_teacher_mappings(request):
    classes = Class.objects.all()
    subjects = Subject.objects.all()

    data = []
    for class_obj in classes:
        row = {}
        row['class'] = class_obj.name
        subject_list = []
        for subject in subjects:
            # Get all relations for this class and subject
            relations = Teacher_Subject_Class_Relation.objects.filter(
                assigned_classes=class_obj, subject=subject
            )
            if relations.exists():
                # Join teacher names if more than one relation exists
                teacher_names = ", ".join([relation.teacher.name for relation in relations if relation.teacher])
            else:
                teacher_names = "Empty"
            subject_list.append({"subject": subject.name, "teacher": teacher_names})
        row['subjects'] = subject_list
        data.append(row)
    return JsonResponse(data, safe=False)

def create_class(request):
    """Create a new class from frontend"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            class_name = data.get("name").strip()

            if not class_name:
                return JsonResponse({"success": False, "message": "Class name cannot be empty!"})

            new_class, created = Class.objects.get_or_create(name=class_name)

            if created:
                return JsonResponse({"success": True, "message": f"Class '{new_class.name}' created successfully!", "id": new_class.id})
            else:
                return JsonResponse({"success": False, "message": "Class already exists!"})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=400)

# Create a new Subject
def create_subject(request):
    if request.method == "POST":
        name = request.POST.get("name")
        if Subject.objects.filter(name=name).exists():
            return JsonResponse({"success": False, "message": "Subject already exists!"})
        Subject.objects.create(name=name)
        return JsonResponse({"success": True, "message": "Subject created successfully!"})
    return JsonResponse({"success": False, "message": "Invalid request method!"})


from .models import Subject, SubSubject

def create_sub_subject(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get("name")
            subject_id = data.get("subject_id")  # Get parent subject ID

            # Ensure the parent subject exists
            try:
                subject = Subject.objects.get(id=subject_id)
            except Subject.DoesNotExist:
                return JsonResponse({"success": False, "message": "Parent subject not found!"})

            # Check if the sub-subject already exists
            if SubSubject.objects.filter(name=name, subject=subject).exists():
                return JsonResponse({"success": False, "message": "Sub-subject already exists!"})

            # Create the sub-subject
            SubSubject.objects.create(name=name, subject=subject)
            return JsonResponse({"success": True, "message": "Sub-subject created successfully!"})

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error: {str(e)}"})

    return JsonResponse({"success": False, "message": "Invalid request method!"})
# views.py
import json
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Class, SubSubject, ClassSubSubjectAssignment



@csrf_exempt  # Only use this if CSRF issues occur, remove in production
def create_class(request):
    """Create a new class from frontend"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            class_name = data.get("name", "").strip()

            if not class_name:
                return JsonResponse({"success": False, "message": "Class name cannot be empty!"}, status=400)

            # Check if class already exists
            new_class, created = Class.objects.get_or_create(name=class_name)

            if created:
                return JsonResponse({"success": True, "message": f"Class '{new_class.name}' created successfully!", "id": new_class.id})
            else:
                return JsonResponse({"success": False, "message": "Class already exists!"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON data"}, status=400)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=405)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Subject

@csrf_exempt  # Use only for debugging; better to use CSRF tokens
def create_subject(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            subject_name = data.get("name", "").strip()

            if not subject_name:
                return JsonResponse({"success": False, "message": "Subject name cannot be empty!"}, status=400)

            new_subject, created = Subject.objects.get_or_create(name=subject_name)

            if created:
                return JsonResponse({"success": True, "message": f"Subject '{new_subject.name}' created successfully!"})
            else:
                return JsonResponse({"success": False, "message": "Subject already exists!"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON data"}, status=400)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=405)

from django.shortcuts import render
from django.http import JsonResponse
from .models import YearlyEvent
from datetime import datetime

def get_yearly_events(request):
    year = request.GET.get("year", datetime.now().year)
    events = YearlyEvent.objects.filter(start_date__year=year)
    
    event_list = [
        {
            "title": event.title,
            "category": event.category,
            "start_date": event.start_date.strftime("%Y-%m-%d"),
            "end_date": event.end_date.strftime("%Y-%m-%d"),
        }
        for event in events
    ]

    return JsonResponse(event_list, safe=False)


from django.http import JsonResponse
from django.shortcuts import render
from .models import Class, Teacher_Subject_Class_Relation, Timetable


def get_teachers_subjects(request):
    teacher_subjects = list(Teacher_Subject_Class_Relation.objects.values("id", "teacher__name", "subject__name"))
    formatted_data = [{"teacher": ts["teacher__name"], "subject": ts["subject__name"]} for ts in teacher_subjects]
    return JsonResponse(formatted_data, safe=False)

from django.shortcuts import render
from .models import YearlyEvent
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date

def year_planner_view(request):
    """Render the Year Planner page"""
    return render(request, "school_management/yearly_planner.html")

def get_events(request):
    """Fetch all events to display in FullCalendar"""
    events = YearlyEvent.objects.all()
    data = [
        {
            "id": event.id,
            "title": event.title,
            "start": event.start_date.strftime("%Y-%m-%d"),
            "end": event.end_date.strftime("%Y-%m-%d"),
        }
        for event in events
    ]
    return JsonResponse(data, safe=False)

@csrf_exempt
def add_event(request):
    """Create a new event via AJAX"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            event = YearlyEvent.objects.create(
                title=data["title"],
                category=data.get("category", "Other"),
                start_date=parse_date(data["start"]),
                end_date=parse_date(data["end"]),
            )
            return JsonResponse({"success": True, "message": "Event added successfully!", "id": event.id})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

@csrf_exempt
def update_event(request):
    """Update an existing event title"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            event = YearlyEvent.objects.get(id=data["id"])
            event.title = data["title"]
            event.save()
            return JsonResponse({"success": True, "message": "Event updated successfully!"})
        except YearlyEvent.DoesNotExist:
            return JsonResponse({"success": False, "error": "Event not found"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)



from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Teacher
from django.db.models import Q


def check_teacher_exists(request):
    name = request.GET.get("name", "").strip().lower()
    phone = request.GET.get("phone", "").strip()

    if not name or not phone:
        return JsonResponse({"error": "Missing required parameters"}, status=400)

    exists = Teacher.objects.filter(name__iexact=name, phone=phone).exists()
    
    return JsonResponse({"exists": exists})


from django.shortcuts import get_object_or_404
from django.http import JsonResponse
import json
from .models import Class, Subject, Teacher_Subject_Class_Relation, Teacher
@csrf_exempt  # Use only for debugging; better to use CSRF tokens
def create_subject(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            subject_name = data.get("name", "").strip()

            if not subject_name:
                return JsonResponse({"success": False, "message": "Subject name cannot be empty!"}, status=400)

            new_subject, created = Subject.objects.get_or_create(name=subject_name)

            if created:
                return JsonResponse({"success": True, "message": f"Subject '{new_subject.name}' created successfully!"})
            else:
                return JsonResponse({"success": False, "message": "Subject already exists!"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON data"}, status=400)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=405)


def assign_subjects_to_class(request):
    """Assign selected subjects to a class (with or without a teacher)."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            class_id = data.get("class_id")
            subject_ids = data.get("subject_ids", [])
            teacher_id = data.get("teacher_id", None)  # ✅ Allow NULL teacher

            assigned_class = get_object_or_404(Class, id=class_id)
            teacher = get_object_or_404(Teacher, id=teacher_id) if teacher_id else None  # ✅ Allow optional teacher

            for subject_id in subject_ids:
                subject = get_object_or_404(Subject, id=subject_id)
                relation, created = Teacher_Subject_Class_Relation.objects.get_or_create(
                    teacher=teacher, subject=subject  # ✅ Works even if teacher is NULL
                )
                relation.assigned_classes.add(assigned_class)

            return JsonResponse({"message": "Subjects assigned successfully!"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request method!"}, status=400)


def get_assigned_subjects(request, class_id):
    """Fetch Subjects Already Assigned to a Class"""
    assigned_class = get_object_or_404(Class, id=class_id)
    assigned_subjects = list(assigned_class.subjects.values("id", "name"))
    return JsonResponse(assigned_subjects, safe=False)
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import Class, Subject
import json
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
import json
from .models import Class, Subject, SubSubject, Teacher_Subject_Class_Relation
from django.views.decorators.csrf import csrf_exempt

from django.shortcuts import render
from .models import Class, SubSubject

def subsubject_management_pg(request):
    """
    Render the page for managing sub-subjects.
    This page contains a dropdown of classes and a list of sub-subjects.
    """
    classes = Class.objects.all().order_by("name")  
    subsubjects = SubSubject.objects.all().order_by("name")  # Fetch all sub-subjects

    return render(request, "school_management/subsubject_management.html", {
        "classes": classes,
        "subsubjects": subsubjects,  # Pass sub-subjects to the template
    })

    
def get_subjects_by_class(request, class_id):
    """
    Return subjects assigned to the given class.
    This example assumes subjects are linked to classes via Teacher_Subject_Class_Relation.
    Adjust as needed if your subject assignment logic is different.
    """
    relations = Teacher_Subject_Class_Relation.objects.filter(assigned_classes__id=class_id).select_related("subject")
    # Use a set to ensure uniqueness
    subjects_set = {relation.subject for relation in relations}
    subjects = [{"id": s.id, "name": s.name} for s in subjects_set]
    return JsonResponse(subjects, safe=False)

def get_subsubjects_by_subject(request, subject_id):
    """
    Return the list of sub-subjects for the given subject.
    """
    subs = SubSubject.objects.filter(subject_id=subject_id)
    sub_list = [{"id": s.id, "name": s.name} for s in subs]
    return JsonResponse(sub_list, safe=False)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import get_object_or_404
from .models import Subject, SubSubject, Class, ClassSubSubjectAssignment

@csrf_exempt
def create_sub_subject(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get("name", "").strip()
            subject_id = data.get("subject_id")
            class_id = data.get("class_id")  # Optional: Assign sub-subject to a class

            if not name or not subject_id:
                return JsonResponse({"success": False, "message": "Sub-subject name and subject ID are required!"}, status=400)

            # Get Subject
            subject = get_object_or_404(Subject, id=subject_id)

            # Check if Sub-Subject exists
            subsubject, created = SubSubject.objects.get_or_create(name=name, subject=subject)

            if not created:
                return JsonResponse({"success": False, "message": "Sub-subject already exists!"}, status=400)

            # Assign to Class if `class_id` is provided
            if class_id:
                try:
                    school_class = Class.objects.get(id=class_id)
                    ClassSubSubjectAssignment.objects.get_or_create(school_class=school_class, subject=subject, subsubject=subsubject)
                except Class.DoesNotExist:
                    return JsonResponse({"success": False, "message": "Class not found!"}, status=404)

            return JsonResponse({
                "success": True,
                "message": f"Sub-subject '{subsubject.name}' created successfully!",
                "subject": subject.name,
                "assigned_to_class": class_id if class_id else "No class assigned"
            })

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON data!"}, status=400)
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)

    return JsonResponse({"success": False, "message": "Invalid request method!"}, status=405)

from django.shortcuts import render
from .models import Class, SubSubject, Teacher_Subject_Class_Relation

def list_class_subjects_subsubjects(request):
    """
    Build a nested data structure that contains:
      - Each class
      - For each class, all subjects assigned (via Teacher_Subject_Class_Relation)
      - For each subject, all sub‑subjects from SubSubject model
    """
    class_data = []
    
    classes = Class.objects.all().order_by('name')
    for cls in classes:
        # Get subjects assigned to this class from Teacher_Subject_Class_Relation
        relations = Teacher_Subject_Class_Relation.objects.filter(assigned_classes=cls).select_related('subject')
        # Use a set to ensure unique subjects
        subjects_set = {relation.subject for relation in relations}
        
        subjects_list = []
        for subject in subjects_set:
            # Get sub-subjects for this subject
            subsubjects = SubSubject.objects.filter(subject=subject)
            subjects_list.append({
                'subject': subject,
                'subsubjects': subsubjects
            })
        
        class_data.append({
            'class': cls,
            'subjects': subjects_list
        })
    
    return render(request, "school_management/class_subject_subsubject_list.html", {"class_data": class_data})
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import SubSubject

@csrf_exempt
def delete_subsubject(request, subsubject_id):
    """
    Delete a sub-subject via AJAX.
    """
    try:
        subsubject = SubSubject.objects.get(id=subsubject_id)
        subsubject.delete()
        return JsonResponse({"success": True})
    except SubSubject.DoesNotExist:
        return JsonResponse({"success": False, "error": "Sub-subject not found"})


from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from .models import Exam, Class, Subject, SubSubject, ExamSchedule, InternalAssessment

def internal_assessment_view(request):
    exams = Exam.objects.all().order_by("name")
    selected_exam_id = request.GET.get("exam_id")
    classes = Class.objects.all().order_by("name")
    columns = []
    flat_columns = []
    exam = None

    # If exam is selected, fetch related subjects and sub-subjects
    if selected_exam_id:
        try:
            exam = Exam.objects.get(id=selected_exam_id)
        except (Exam.DoesNotExist, ValueError):
            exam = None

        if exam:
            exam_schedules = ExamSchedule.objects.filter(exam=exam).select_related("subject")
            subject_ids = exam_schedules.values_list("subject_id", flat=True)
            subjects = Subject.objects.filter(id__in=subject_ids).order_by("name")
            subject_map = {s.name: s for s in subjects}

            for subject in subjects:
                subsubjects = subject.sub_subjects.all().order_by("name")
                if subsubjects.exists():
                    sub_list = [(sub.id, sub.name) for sub in subsubjects]
                    columns.append((subject.name, sub_list))

            for subj_name, sub_list in columns:
                subject = subject_map[subj_name]
                for sub_id, sub_name in sub_list:
                    flat_columns.append({
                        "subject_id": subject.id,
                        "subject": subj_name,
                        "sub_id": sub_id,
                        "sub": sub_name
                    })

    # Handle POST request to save marks
    if request.method == "POST" and exam:
        for cls in classes:
            for col in flat_columns:
                input_name = f"mark_{cls.id}_{col['subject_id']}_{col['sub_id']}"
                mark_value = request.POST.get(input_name)
                if mark_value is not None and mark_value.strip() != "":
                    assessment, _ = InternalAssessment.objects.get_or_create(
                        school_class=cls,
                        exam=exam,
                        subject_id=col["subject_id"],
                        subsubject_id=col["sub_id"],
                        defaults={'max_marks': 100}  # Set default max_marks or customize as needed
                    )
                    assessment.obtained_marks = mark_value
                    assessment.save()
        messages.success(request, "Marks saved successfully.")
        return redirect(f"{reverse('internal_assessment')}?exam_id={selected_exam_id}")

    # Build table data for display
    table_data = []
    for cls in classes:
        marks_dict = {
            (col["subject_id"], col["sub_id"]): {"mark": "-", "max": "-"}
            for col in flat_columns
        }

        if exam:
            assessments = InternalAssessment.objects.filter(school_class=cls, exam=exam)
            for assessment in assessments:
                key = (
                    assessment.subject.id,
                    assessment.subsubject.id if assessment.subsubject else None
                )
                if key in marks_dict:
                    marks_dict[key] = {
                        "mark": assessment.obtained_marks if assessment.obtained_marks is not None else "-",
                        "max": assessment.max_marks if assessment.max_marks is not None else "-"
                    }

        flat_marks = [
            {
                "subject": col["subject"],
                "subject_id": col["subject_id"],
                "sub": col["sub"],
                "sub_id": col["sub_id"],
                "mark": marks_dict.get((col["subject_id"], col["sub_id"]), {}).get("mark", "-"),
                "max": marks_dict.get((col["subject_id"], col["sub_id"]), {}).get("max", "-"),
            }
            for col in flat_columns
        ]

        table_data.append({
            "class_id": cls.id,
            "class_name": cls.name,
            "marks": flat_marks
        })

    # Create structure for column headers in template
    display_columns = [(subj_name, [sub_name for (_, sub_name) in sub_list]) for subj_name, sub_list in columns]

    context = {
        "exams": exams,
        "selected_exam_id": int(selected_exam_id) if selected_exam_id else None,
        "columns": display_columns,
        "table_data": table_data,
    }
    return render(request, "school_management/Exam_dep_internal_assment.html", context)
import json
from .models import InternalAssessment, Student, Subject, SubSubject, Class, Exam


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import InternalAssessment
@csrf_exempt
def update_internal_marks(request):
    if request.method == "POST":
        exam_id = request.POST.get("exam_id")
        success_count = 0
        fail_count = 0
        debug_log = []

        if not exam_id:
            return JsonResponse({"success": False, "message": "Missing exam_id in request."})

        for key, value in request.POST.items():
            if key.startswith("mark_"):
                try:
                    _, class_id, subject_id, sub_id = key.split("_")
                    sub_id = None if sub_id == "None" else sub_id
                    mark_value = int(value.strip()) if value.strip().isdigit() else 0

                    # Ensure the record exists before updating
                    record, created = InternalAssessment.objects.update_or_create(
                        exam_id=exam_id,
                        school_class_id=class_id,
                        subject_id=subject_id,
                        subsubject_id=sub_id,
                        defaults={"max_marks": mark_value}
                    )

                    debug_log.append(f"{'Created' if created else 'Updated'} record: {record.id} => max_marks={mark_value}")
                    success_count += 1
                except Exception as e:
                    debug_log.append(f"Error processing {key}: {str(e)}")
                    fail_count += 1

        return JsonResponse({
            "success": True,
            "message": f"{success_count} marks updated, {fail_count} failed.",
            "log": debug_log
        })

    return JsonResponse({"success": False, "message": "Invalid request method."})


from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect
from django.contrib import messages

from .models import  InternalAssessment, Exam, Subject, SubSubject
from django.contrib import messages



from .models import Exam  # make sure this model is imported
from django.shortcuts import render, get_object_or_404
from .models import Class, Exam, Student, ClassSubSubjectAssignment, InternalAssessment

def internal_assessment_entry(request):
    class_list = Class.objects.all()
    exam_list = Exam.objects.all()

    selected_class_id = request.GET.get('class_id')
    selected_exam_id = request.GET.get('exam_id')

    print("Selected Class ID:", selected_class_id)
    print("Selected Exam ID:", selected_exam_id)

    context = {
        'class_list': class_list,
        'exam_list': exam_list,
        'selected_class_id': selected_class_id,
        'selected_exam_id': selected_exam_id,
    }

    if selected_class_id and selected_exam_id:
        try:
            selected_class = Class.objects.get(id=selected_class_id)
            selected_exam = Exam.objects.get(id=selected_exam_id)
            print("Selected Class:", selected_class)
        except Class.DoesNotExist:
            print("Class does not exist")
            selected_class = None

        if selected_class:
            students = Student.objects.filter(assigned_class=selected_class)
            print("Students:", students)
            print("Student count:", students.count())

            assignments = ClassSubSubjectAssignment.objects.filter(school_class=selected_class)
            print("Assignments:", assignments)
            print("Assignment count:", assignments.count())

            subject_structure = {}

            for assign in assignments:
                subject = assign.subject
                sub = assign.subsubject

                # Fetch max_marks from InternalAssessment
                assessment = InternalAssessment.objects.filter(
                    school_class=selected_class,
                    exam=selected_exam,
                    subject=subject,
                    subsubject=sub
                ).first()

                max_marks = assessment.max_marks if assessment else "—"

                if subject not in subject_structure:
                    subject_structure[subject] = {
                        'subject': subject.name,
                        'subject_id': subject.id,
                        'subsubjects': []
                    }

                subject_structure[subject]['subsubjects'].append({
                    'id': sub.id,
                    'name': sub.name,
                    'max_marks': max_marks
                })

            context['processed_students'] = students
            context['subject_structure'] = list(subject_structure.values())

    return render(request, 'school_management/teacher_Internal_entry.html', context)


@require_POST
@csrf_exempt
def save_internal_marks_by_class(request):
    try:
        for key, value in request.POST.items():
            if key.startswith("mark_") and value.strip() != "":
                parts = key.split("_")  # Expected: ['mark', student_id, subject_id, subsubject_id]
                if len(parts) == 4:
                    student_id, subject_id, subsubject_id = parts[1], parts[2], parts[3]
                    try:
                        marks = int(value)
                    except ValueError:
                        marks = 0  # or skip if you want

                    InternalAssessment.objects.update_or_create(
                        student_id=student_id,
                        subject_id=subject_id,
                        subsubject_id=subsubject_id,
                        defaults={"marks": marks}
                    )
        messages.success(request, "Internal assessment marks saved successfully.")
    except Exception as e:
        messages.error(request, f"Error saving marks: {e}")
    return redirect("internal_assessment_entry")


def student_assessment_view(request, class_id):
    students = Student.objects.filter(school_class_id=class_id)
    assessments = InternalAssessment.objects.filter(school_class_id=class_id)

    subjects = {}
    for assessment in assessments:
        subject_name = assessment.subject.name
        if subject_name not in subjects:
            subjects[subject_name] = []
        subjects[subject_name].append(assessment)

    context = {
        "students": students,
        "subjects": subjects,
    }
    return render(request, "assessment_page.html", context)


def assign_student_role(request):
    """Assign leadership roles to students"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            student_id = data.get("student_id")
            role = data.get("role")

            student = get_object_or_404(Student, id=student_id)
            student.role = role
            student.save()

            return JsonResponse({"message": f"Role '{student.get_role_display()}' assigned to {student.name}."})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request method!"}, status=400)
from django.utils.dateparse import parse_date

def create_exam(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            exam_name = data.get("name", "").strip()
            exam_date_str = data.get("exam_date", "").strip()  # ✅ Ensure it's a string

            if not exam_name or not exam_date_str:
                return JsonResponse({"error": "Exam name and date are required"}, status=400)

            exam_date = parse_date(exam_date_str)  # ✅ Convert string to date
            if not exam_date:
                return JsonResponse({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

            exam = Exam.objects.create(name=exam_name, exam_date=exam_date)
            return JsonResponse({"message": "Exam created successfully!"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)


def get_teacher_availability(request):
    teachers = Teacher.objects.all()
    availability = {teacher.name: list(teacher.assigned_classes.values_list('name', flat=True)) for teacher in teachers}
    return JsonResponse(availability)

def get_predefined_plans(request):
    plans = {t.class_obj.id: list(t.class_obj.subjects.values_list('name', flat=True)) for t in Timetable.objects.all()}
    return JsonResponse(plans)


from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from .models import Timetable, BreakPeriod


def get_teacher_workload(request, teacher_name):
    teacher = Teacher.objects.filter(name=teacher_name).first()
    if not teacher:
        return JsonResponse({"error": "Teacher not found"}, status=404)

    periods = Timetable.objects.filter(teacher_subject__teacher=teacher).count()
    return JsonResponse({"teacher": teacher_name, "periods": periods})
def get_breaks(request):
    breaks = BreakPeriod.objects.all()
    data = [
        {
            "day": b.get_day_display(),
            "period": b.period,
            "start_time": b.start_time.strftime("%H:%M"),
            "end_time": b.end_time.strftime("%H:%M"),
            "break_type": b.get_break_type_display()
        }
        for b in breaks
    ]
    return JsonResponse(data, safe=False)



def get_entities(request, timetable_type):
    """Fetch classes, teachers, or the principal for timetable selection."""
    if timetable_type == 'class':
        entities = [{'id': c.id, 'name': c.name} for c in Class.objects.all()]  # Removed subjects__isnull=False
    elif timetable_type == 'teacher':
        entities = [{'id': t.id, 'name': t.name} for t in Teacher.objects.all()]
    elif timetable_type == 'principal':
        entities = [{'id': 1, 'name': "Principal"}]  # Static entry for Principal's timetable
    else:
        return JsonResponse({'error': 'Invalid timetable type'}, status=400)
    return JsonResponse({'entities': entities})

def daywise_timetable_entry_view(request):
    """Render the Day-wise Timetable Entry page."""
    classes = Class.objects.all()
    teachers_subjects = Teacher_Subject_Class_Relation.objects.select_related('teacher', 'subject').all()
    return render(request, "school_management/daywise_timetable_entry.html", {
        "classes": classes,
        "teachers_subjects": teachers_subjects
    })


def check_timetable_conflicts(request):
    """Check for scheduling conflicts before saving timetable."""
    if request.method == 'POST':
        data = json.loads(request.body)
        timetable = data.get("timetable", [])
        conflict_found = False
        conflict_message = ""
        for entry in timetable:
            teacher_id = entry["teacher_id"]
            period = entry["period"]
            day = entry.get("day")
            
            existing_entry = Timetable.objects.filter(
                teacher_subject__teacher_id=teacher_id, 
                period=period, 
                day=day
            ).exists()
            
            if existing_entry:
                conflict_found = True
                teacher = get_object_or_404(Teacher, id=teacher_id)
                conflict_message = f"Conflict: {teacher.name} is already assigned in this period."
                break
        
        return JsonResponse({"conflict": conflict_found, "message": conflict_message})
    
    return JsonResponse({"error": "Invalid request"}, status=400)

def update_timetable(request):
    """Update timetable entries while handling conflicts and teacher substitution."""
    if request.method == 'POST':
        data = json.loads(request.body)
        timetable_entries = data.get("timetable", [])

        for entry in timetable_entries:
            class_id = entry.get("class_id")
            teacher_id = entry.get("teacher_id")
            subject_id = entry.get("subject_id")
            period = entry.get("period")
            day = entry.get("day")

            # Check for conflicts before saving
            if Timetable.objects.filter(teacher_subject__teacher_id=teacher_id, period=period, day=day).exists():
                return JsonResponse({"error": "Conflict detected! Teacher is already assigned."}, status=400)

            teacher_subject = get_object_or_404(Teacher_Subject_Class_Relation, teacher_id=teacher_id, subject_id=subject_id)
            class_obj = get_object_or_404(Class, id=class_id)

            Timetable.objects.update_or_create(
                class_obj=class_obj, period=period, day=day,
                defaults={"teacher_subject": teacher_subject}
            )

        return JsonResponse({"message": "Timetable updated successfully!"})
    
    return JsonResponse({"error": "Invalid request"}, status=400)


def get_teachers_subjects_sorted(request, class_id):
    """Fetch assigned teachers and subjects for a class."""
    assigned_class = get_object_or_404(Class, id=class_id)
    teacher_subjects = Teacher_Subject_Class_Relation.objects.filter(assigned_classes=assigned_class).select_related('teacher', 'subject')

    teachers = {ts.teacher.id: ts.teacher.name for ts in teacher_subjects}
    subjects = {ts.subject.id: ts.subject.name for ts in teacher_subjects}

    return JsonResponse({"teachers": list(teachers.items()), "subjects": list(subjects.items())})

import json
from django.shortcuts import render
from .models import Class, Timetable

def timetable_view(request):
    """Render the timetable management page with class list and saved data"""
    classes = Class.objects.all()
    periods = list(range(1, 9))  # Periods 1 to 8

    # Convert DAY_CHOICES tuples to a list of strings (like 'Monday', 'Tuesday', etc.)
    days = [d[0] for d in Timetable.DAY_CHOICES]

    saved_entries = Timetable.objects.all()  # Get all saved timetable entries

    # Structure the timetable data as a dictionary for easy access
    timetable_data = {}
    for entry in saved_entries:
        # Using a string as a key instead of a tuple
        key = f"{entry.class_obj.id}_{entry.period}"  # Concatenate class ID and period as a string
        if key not in timetable_data:
            timetable_data[key] = {}
        
        # Access teacher and subject details through the teacher_subject relationship
        if entry.teacher_subject:
            teacher = entry.teacher_subject.teacher
            subject = entry.teacher_subject.subject
            timetable_data[key] = {
                'teacher_id': teacher.id,
                'subject_id': subject.id,
                'teacher_name': teacher.name,
                'subject_name': subject.name,
            }
        else:
            timetable_data[key] = {
                'teacher_id': None,
                'subject_id': None,
                'teacher_name': 'Not Assigned',
                'subject_name': 'Not Assigned',
            }

    # Set default period to 8
    default_period = 8

    return render(request, 'school_management/timetable_management.html', {
        'classes': classes,
        'periods': periods,
        'days': days,  # Pass cleaned up day list
        'timetable_data': timetable_data,
        'default_period': default_period
    })

from django.http import JsonResponse


from collections import defaultdict
from django.http import JsonResponse
from .models import Timetable, Teacher, Subject
from django.db import transaction
def save_timetable(request):
    if request.method == 'POST':
        try:
            # Access the 'valid_rows' data from POST request (as JSON from hidden textarea)
            valid_rows_json = request.POST.get('valid_rows')
            if not valid_rows_json:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No valid_rows data found in the request.'
                })

            # Parse the valid_rows JSON data
            try:
                valid_rows = json.loads(valid_rows_json)
            except json.JSONDecodeError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid JSON format in request body. Please try again.'
                })

            valid_rows_list = []
            class_ids_to_clear = set()

            # Loop through each row of valid data
            for row in valid_rows:
                day_name = row.get('day_name')
                class_id = row.get('class_id')
                teacher_id = row.get('teacher_id')
                subject_id = row.get('subject_id')
                period_number = row.get('period_number')  # This will be used directly

                # Ensure all necessary fields are present
                if not all([day_name, class_id, teacher_id, subject_id, period_number]):
                    print(f"Skipping row due to missing fields: {row}")
                    continue

                # Fetch Teacher and Subject objects
                try:
                    teacher = Teacher.objects.get(id=teacher_id)
                except Teacher.DoesNotExist:
                    print(f"Invalid teacher_id: {teacher_id}")
                    continue

                try:
                    subject = Subject.objects.get(id=subject_id)
                except Subject.DoesNotExist:
                    print(f"Invalid subject_id: {subject_id}")
                    continue

                valid_rows_list.append({
                    'day_name': day_name,
                    'class_id': class_id,
                    'teacher': teacher,
                    'subject': subject,
                    'period_number': period_number  # Save period number from form
                })
                class_ids_to_clear.add(class_id)

            if not valid_rows_list:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No valid rows to save. Please ensure all fields are correctly filled.'
                })

            # Delete old timetable entries for the same classes
            Timetable.objects.filter(class_obj_id__in=class_ids_to_clear).delete()

            timetable_objects = []

            # Loop to create timetable entries
            for row in valid_rows_list:
                timetable_objects.append(Timetable(
                    class_obj_id=row['class_id'],
                    day=row['day_name'],
                    period=row['period_number'],  # Use period number from the form
                    teacher=row['teacher'],
                    subject=row['subject']
                ))

            # Bulk create timetable entries using transaction for atomic operation
            with transaction.atomic():
                Timetable.objects.bulk_create(timetable_objects)

            return JsonResponse({
                'status': 'success',
                'message': f'{len(timetable_objects)} timetable entries successfully saved!'
            })

        except Exception as e:
            # Handle unexpected errors
            return JsonResponse({
                'status': 'error',
                'message': f'Unexpected error: {str(e)}'
            })

    else:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid request method. Please use POST.'
        })


from django.views.decorators.http import require_GET
@require_GET
def get_saved_timetable(request, day):
    """
    Returns the saved timetable entries for the given day with detailed info.
    """
    try:
        timetable_entries = Timetable.objects.select_related(
            'class_obj', 'teacher_subject__teacher', 'teacher_subject__subject'
        ).filter(day__iexact=day)

        data = []

        for entry in timetable_entries:
            if entry.is_break:
                data.append({
                    "class_id": entry.class_obj.id,
                    "class_name": entry.class_obj.name,
                    "period": entry.period,
                    "is_break": True,
                    "start_time": entry.start_time.strftime("%H:%M") if entry.start_time else None,
                    "end_time": entry.end_time.strftime("%H:%M") if entry.end_time else None,
                    "teacher_name": None,
                    "subject_name": None,
                })
            else:
                data.append({
                    "class_id": entry.class_obj.id,
                    "class_name": entry.class_obj.name,
                    "period": entry.period,
                    "is_break": False,
                    "start_time": entry.start_time.strftime("%H:%M") if entry.start_time else None,
                    "end_time": entry.end_time.strftime("%H:%M") if entry.end_time else None,
                    "teacher_id": entry.teacher_subject.teacher.id if entry.teacher_subject else None,
                    "teacher_name": entry.teacher_subject.teacher.name if entry.teacher_subject else "Unassigned",
                    "subject_id": entry.teacher_subject.subject.id if entry.teacher_subject else None,
                    "subject_name": entry.teacher_subject.subject.name if entry.teacher_subject else "Unassigned",
                })

        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def get_timetable(request, class_id):
    """Fetch timetable data for a class"""
    timetable = Timetable.objects.filter(class_obj_id=class_id).select_related("teacher_subject__teacher", "teacher_subject__subject")
    data = [
        {
            "period": t.period,
            "day": t.day,
            "teacher": t.teacher_subject.teacher.name if t.teacher_subject else "",
            "subject": t.teacher_subject.subject.name if t.teacher_subject else "",
            "is_break": t.is_break,
            "start_time": t.start_time.strftime("%H:%M") if t.start_time else "",
            "end_time": t.end_time.strftime("%H:%M") if t.end_time else "",
        }
        for t in timetable
    ]
    return JsonResponse(data, safe=False)


@csrf_exempt
def update_timetable(request):
    """Update timetable entries while handling conflicts"""
    if request.method == "POST":
        data = json.loads(request.body)
        timetable_entries = data.get("timetable", [])

        for entry in timetable_entries:
            class_id = entry.get("class_id")
            teacher_id = entry.get("teacher_id")
            subject_id = entry.get("subject_id")
            period = entry.get("period")
            day = entry.get("day")
            is_break = entry.get("is_break", False)

            if not is_break:
                teacher_subject = get_object_or_404(
                    Teacher_Subject_Class_Relation,
                    teacher_id=teacher_id,
                    subject_id=subject_id,
                )
            else:
                teacher_subject = None

            class_obj = get_object_or_404(Class, id=class_id)
            Timetable.objects.update_or_create(
                class_obj=class_obj, period=period, day=day,
                defaults={
                    "teacher_subject": teacher_subject,
                    "is_break": is_break,
                    "start_time": entry.get("start_time"),
                    "end_time": entry.get("end_time"),
                }
            )

        return JsonResponse({"message": "Timetable updated successfully!"})
    
    return JsonResponse({"error": "Invalid request"}, status=400)
def get_classes(request):
    """Fetch all classes from the database"""
    classes = list(Class.objects.values("id", "name"))
    return JsonResponse(classes, safe=False)

def get_teachers(request):
    """Fetch all teachers as JSON"""
    teachers = list(Teacher.objects.values("id", "name"))
    return JsonResponse(teachers, safe=False)

def get_subjects(request):
    """Fetch all subjects as JSON"""
    subjects = list(Subject.objects.values("id", "name"))
    return JsonResponse(subjects, safe=False)
def get_available_teachers(request, period):
    """Fetch available teachers for a specific period, excluding already selected ones"""
    selected_teachers = request.GET.getlist('selected_teachers[]')  # Get selected teachers from frontend
    
    available_teachers = Teacher.objects.exclude(id__in=selected_teachers).values("id", "name")
    return JsonResponse(list(available_teachers), safe=False)
def get_teachers_by_class(request, class_id):
    """Fetch teachers assigned to a specific class"""
    teachers = Teacher_Subject_Class_Relation.objects.filter(assigned_classes__id=class_id)
    teachers_list = list(teachers.values("teacher__id", "teacher__name").distinct())
    formatted_teachers = [{"id": t["teacher__id"], "name": t["teacher__name"]} for t in teachers_list]
    return JsonResponse(formatted_teachers, safe=False)

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Teacher, Teacher_Subject_Class_Relation, Subject

def get_subjects_by_teacher(request, teacher_id):
    try:
        teacher = get_object_or_404(Teacher, id=teacher_id)

        # First try to get subjects from Teacher_Subject_Class_Relation
        subjects = Teacher_Subject_Class_Relation.objects.filter(teacher=teacher).values("subject__id", "subject__name").distinct()
        subjects_list = [{"id": s["subject__id"], "name": s["subject__name"]} for s in subjects]

        # If no subjects found in the relation table, fall back to the teacher's direct M2M field
        if not subjects_list:
            subjects = teacher.subjects.all()
            subjects_list = [{"id": s.id, "name": s.name} for s in subjects]

        return JsonResponse(subjects_list, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Teacher, Teacher_Subject_Class_Relation, Class

def get_subjects_by_teacher_and_class(request, teacher_id, class_id=None):
    try:
        teacher = get_object_or_404(Teacher, id=teacher_id)

        # Start with relation-based subject filtering
        relations = Teacher_Subject_Class_Relation.objects.filter(teacher=teacher)

        if class_id:
            relations = relations.filter(assigned_classes__id=class_id)

        subjects = relations.values("subject__id", "subject__name").distinct()
        subjects_list = [{"id": s["subject__id"], "name": s["subject__name"]} for s in subjects]

        # Fallback: M2M field if relation doesn't exist
        if not subjects_list:
            if class_id:
                # Optionally, return only if class is part of teacher's classes via M2M logic
                class_obj = get_object_or_404(Class, id=class_id)
                if class_obj in teacher.classes.all():
                    subjects = teacher.subjects.all()
                    subjects_list = [{"id": s.id, "name": s.name} for s in subjects]
            else:
                subjects = teacher.subjects.all()
                subjects_list = [{"id": s.id, "name": s.name} for s in subjects]

        return JsonResponse(subjects_list, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

import json
from django.shortcuts import get_object_or_404
from .models import Timetable, Class, Teacher_Subject_Class_Relation  # Assuming models are in the same app

def teacher_timetable_pg(request):
    """Render the teacher timetable selection page with short periods."""
    teachers = Teacher.objects.all()
    periods = [f"P{i}" for i in range(1, 10)]  # Generates ["P1", "P2", ..., "P9"]
    return render(request, "school_management/teacher_timetable.html", {"teachers": teachers, "periods": periods})


from django.http import JsonResponse
from .models import Timetable  # Adjust path as needed
import logging

logger = logging.getLogger(__name__)

def get_teacher_timetable(request, teacher_id):
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    periods = list(range(1, 10))  # periods as integers (1 to 9)

    teacher_timetable = []

    for day in days:
        day_data = {"day": day}
        for i in periods:
            try:
                slot = Timetable.objects.filter(teacher_id=teacher_id, day=day, period=i).first()
                period_key = f"P{i}"
                if slot:
                    if getattr(slot, 'is_break', False):
                        day_data[period_key] = {"subject": "Break", "class": "Break"}
                    else:
                        day_data[period_key] = {
                            "subject": getattr(slot.subject, "name", "N/A"),
                            "class": getattr(slot.class_obj, "name", "N/A")
                        }
                else:
                    day_data[period_key] = {"subject": "Free", "class": "Free"}
            except Exception as e:
                logger.error(f"Error fetching period {i} on {day}: {e}")
                day_data[f"P{i}"] = {"subject": "Error", "class": "Error"}
        teacher_timetable.append(day_data)

    return JsonResponse({
        "periods": [str(p) for p in periods],
        "timetable": teacher_timetable
    })

from django.shortcuts import get_object_or_404
from .models import Timetable, Class
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from collections import defaultdict
from .models import Timetable, Class


from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from collections import defaultdict
from .models import Timetable, Class

# views.py
from django.shortcuts import render
from django.http import JsonResponse
from .models import Timetable, Class

def class_timetable_page(request):
    # Fetch all classes to populate the dropdown on the frontend
    classes = Class.objects.all()
    return render(request, "school_management/class_timetable.html", {"classes": classes})

from django.http import JsonResponse
from .models import Timetable

from django.http import JsonResponse
from .models import Timetable

def get_class_timetable(request, class_id):
    from collections import defaultdict

    data = defaultdict(lambda: {f'P{i}': {} for i in range(1, 10)})
    max_period_used = 0

    qs = Timetable.objects.filter(class_obj_id=class_id).order_by('day', 'period')
    for entry in qs:
        period_key = f'P{entry.period}'
        data[entry.day][period_key] = {
            'teacher': entry.teacher.name if entry.teacher else '',
            'subject': entry.subject.name if entry.subject else '',
            'is_break': entry.is_break,
            'start_time': entry.start_time.strftime('%H:%M') if entry.start_time else '',
            'end_time': entry.end_time.strftime('%H:%M') if entry.end_time else '',
        }
        if entry.period > max_period_used:
            max_period_used = entry.period

    timetable_list = []
    for day, periods in data.items():
        row = {'day': day}
        for i in range(1, max_period_used + 1):
            row[f'P{i}'] = periods.get(f'P{i}', {})
        timetable_list.append(row)

    return JsonResponse({
        'timetable': timetable_list,
        'periods': list(range(1, max_period_used + 1))
    })

def free_teachers_view(request):
    """Render the Free Teachers Timetable Page"""
    return render(request, "school_management/free_teachers.html")

from django.http import JsonResponse
from django.shortcuts import render
from .models import Timetable, Teacher, Teacher_Subject_Class_Relation

def get_free_teachers_timetable(request):
    """Fetch free teachers for each period, grouped by all 7 days"""

    # ✅ Get all days and periods
    all_days = [d[0] for d in Timetable.DAY_CHOICES]
    all_periods = sorted([p[0] for p in Timetable.PERIOD_CHOICES])
    all_teachers = set(Teacher.objects.values_list("id", "name"))  # Use ID to ensure uniqueness

    # ✅ Map teacher ID to name
    teacher_id_to_name = {t[0]: t[1] for t in all_teachers}
    teacher_ids = set(teacher_id_to_name.keys())

    # ✅ Initialize the structure
    free_teachers = {
        day: {period: set(teacher_ids) for period in all_periods}
        for day in all_days
    }

    # ✅ Get all assigned teachers (exclude breaks)
    busy_slots = Timetable.objects.filter(is_break=False).select_related("teacher_subject__teacher")

    for entry in busy_slots:
        if entry.teacher_subject and entry.teacher_subject.teacher:
            tid = entry.teacher_subject.teacher.id
            day = entry.day
            period = entry.period

            if day in free_teachers and period in free_teachers[day]:
                free_teachers[day][period].discard(tid)  # remove if assigned

    # ✅ Convert teacher IDs back to names
    free_teachers_named = {
        day: {
            period: [teacher_id_to_name[tid] for tid in sorted(tids)]
            for period, tids in periods.items()
        } for day, periods in free_teachers.items()
    }

    return JsonResponse({
        "periods": all_periods,
        "days": all_days,
        "free_teachers": free_teachers_named
    })

from django.http import JsonResponse
from .models import Timetable, Teacher

def get_free_teachers_timetable(request):
    """Return JSON of free teachers for each period on each day."""

    # Define all periods and all days (from model constants or hardcoded if needed)
    all_periods = [i for i in range(1, 10)]  # Assuming 9 periods
    all_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    # Get all teacher names
    all_teachers = list(Teacher.objects.values_list("name", flat=True))

    # Build free_teachers dictionary
    free_teachers = {}

    for day in all_days:
        free_teachers[day] = {}

        for period in all_periods:
            # Get names of teachers assigned for this period and day
            busy_teachers = Timetable.objects.filter(day=day, period=period).values_list("teacher__name", flat=True)

            # Get free teachers by excluding busy ones
            free = [teacher for teacher in all_teachers if teacher not in busy_teachers]

            free_teachers[day][period] = free

    return JsonResponse({
        "periods": all_periods,
        "days": all_days,
        "free_teachers": free_teachers
    })


from django.shortcuts import render
from .models import Class, Student, Teacher_Subject_Class_Relation

def classroom_overview(request):
    """Fetch all classrooms and their assigned classes with details."""
    
    classes = Class.objects.exclude(id__isnull=True)  # ✅ Exclude classes with no ID
    classroom_data = []

    for cls in classes:
        subject_list = Teacher_Subject_Class_Relation.objects.filter(assigned_classes=cls).values_list("subject__name", flat=True)

        classroom_data.append({
            "id": cls.id,  
            "name": cls.name,
            "class_teacher": cls.class_teacher.name if cls.class_teacher else "No Teacher",  # ✅ Fix: Use class_teacher
            "students_count": Student.objects.filter(assigned_class=cls).count(),
            "subjects": ", ".join(subject_list) if subject_list else "No Subjects",
            "syllabus": cls.syllabus if cls.syllabus else "No syllabus available"
        })

    return render(request, "school_management/classroom_overview.html", {"classroom_data": classroom_data})

from django.shortcuts import render, get_object_or_404
from .models import Class, Student, Teacher_Subject_Class_Relation

def class_detail_view(request, class_id):
    """Fetch details of a single class"""
    
    cls = get_object_or_404(Class, id=class_id)

    # ✅ Fetch subjects from Teacher_Subject_Class_Relation
    subject_list = Teacher_Subject_Class_Relation.objects.filter(assigned_classes=cls).values_list("subject__name", flat=True)

    # ✅ Fetch students of the class
    students = Student.objects.filter(assigned_class=cls)

    class_data = {
        "name": cls.name,
        "class_teacher": cls.class_teacher.name if cls.class_teacher else "No Teacher",
        "class_leader": students.filter(role="Class Leader").first().name if students.filter(role="Class Leader").exists() else "No Leader",
        "students_count": students.count(),
        "subjects": ", ".join(subject_list) if subject_list else "No Subjects",
        "students": students,
        "syllabus": cls.syllabus.url if cls.syllabus else None
    }

    return render(request, "school_management/class_In_detail.html", {"class_data": class_data})

from .models import Teacher_Subject_Class_Relation, Subject, Class

def teacher_subject_class_table(request):
    subjects = Subject.objects.all().order_by('name')
    classes = Class.objects.all().order_by('name')

    # Preload relations
    relations = Teacher_Subject_Class_Relation.objects.select_related('teacher', 'subject').prefetch_related('assigned_classes').all()

    # Create a dictionary to look up relations by subject and class
    relation_map = {}
    for relation in relations:
        for cls in relation.assigned_classes.all():
            relation_map[(relation.subject.id, cls.id)] = relation

    table_data = []

    for subject in subjects:
        row = [subject.name]
        for cls in classes:
            relation = relation_map.get((subject.id, cls.id))
            if relation and relation.teacher:
                # Check if teacher exists before accessing name
                row.append(f"{relation.teacher.name} <br> <small>{relation.teacher.phone}</small>")
            else:
                row.append('<span class="no-data">Not Assigned</span>')
        table_data.append(row)

    return render(request, "school_management/teacher_subject_class_table.html", {
        "classes": classes,
        "table_data": table_data,
    })




from django.shortcuts import render
from .models import YearlyEvent
def event_list(request):
    selected_event_label = request.GET.get('event', None)
    if selected_event_label:
        event = YearlyEvent.objects.filter(name=selected_event_label).first()
    else:
        event = None
    
    events = YearlyEvent.objects.all().order_by('start_date')
    
    return render(request, 'school_management/event_list.html', {
        'events': events,
        'selected_event': event
    })



import json
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import YearlyEvent
from datetime import datetime

def edit_event(request, event_id):
    event = get_object_or_404(YearlyEvent, id=event_id)

    if request.method == 'POST':
        # Get data from the form (using request.POST)
        title = request.POST.get('title')
        category = request.POST.get('category')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        # Basic validation for required fields
        if not title or not category or not start_date or not end_date:
            return JsonResponse({'success': False, 'message': 'All fields are required.'})

        # Convert string dates to actual datetime objects (Django accepts 'YYYY-MM-DD' format)
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format.'})

        # Update the event with new data
        event.title = title
        event.category = category
        event.start_date = start_date
        event.end_date = end_date
        event.save()

        # Redirect to event list page
        return redirect('yearly_event_list')  # or wherever you want to go after editing

    else:
        # Pre-populate the form with the current event details
        context = {
            'event': event
        }
        return render(request, 'school_management/edit_event.html', context)

@csrf_exempt  # Exempting CSRF for the example (But you should ideally handle CSRF)
def delete_event(request, event_id):
    # Check if the request is AJAX and is a POST request
    if request.method == "POST" and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            event = get_object_or_404(YearlyEvent, id=event_id)
            event.delete()
            return JsonResponse({'success': True, 'message': 'Event deleted successfully.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Failed to delete event: {str(e)}'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request.'})


import json  # Make sure this is imported at the top

def principal_dashboard(request):
    """Fetch data and render the Principal Dashboard with an event timeline chart using dates."""

    events = YearlyEvent.objects.all().order_by("start_date")

    event_labels = []
    event_start_dates = []
    event_end_dates = []
    event_ids = []

    if events.exists():
        event_labels = [event.title for event in events]
        event_start_dates = [event.start_date.strftime("%Y-%m-%d") for event in events]
        event_end_dates = [event.end_date.strftime("%Y-%m-%d") for event in events]
        event_ids = [event.id for event in events]  # ✅ Include event IDs here

    return render(request, "school_management/principal_dashboard.html", {
        "event_labels": json.dumps(event_labels),
        "event_start_dates": json.dumps(event_start_dates),
        "event_end_dates": json.dumps(event_end_dates),
        "event_ids": json.dumps(event_ids),  # ✅ Pass event IDs to the template
    })


from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json

from .models import (
    Class,
    Student,
    Subject,
    Exam,
    ExamSchedule,
    Marks,
    StudentSubjectRemark,
    Teacher_Subject_Class_Relation
)

from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Class, Exam, Student, Subject, Teacher_Subject_Class_Relation, ExamSchedule, Marks, StudentSubjectRemark
import json

from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Class, Exam, Student, Subject, Teacher_Subject_Class_Relation, ExamSchedule, Marks, StudentSubjectRemark
import json
from django.shortcuts import render, redirect
from django.contrib import messages

def contact_page(request):
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject', '')
        message = request.POST.get('message')

        # You can handle saving or emailing the message here
        # Example: save to database or send email (not shown here)

        messages.success(request, "Thank you for contacting us!")
        return redirect('contact_page')  # Redirect to clear form

    return render(request, 'school_management/contact.html')

from django.views.decorators.http import require_http_methods
from django.shortcuts import render
import json
@require_http_methods(["GET"])
def teacher_dashboard(request):
    # Fetch GET parameters for class and exam
    class_id = request.GET.get('class_id')
    exam_id = request.GET.get('exam_id')

    # Fetch all classes, exams, students, and subjects
    classes = Class.objects.all()
    exams = Exam.objects.all()
    students = Student.objects.all()
    subjects = Subject.objects.all()

    # Validate and sanitize class_id
    if class_id:
        try:
            class_id = int(class_id)
            if not classes.filter(id=class_id).exists():
                class_id = None
        except ValueError:
            class_id = None

    # Validate and sanitize exam_id
    if exam_id:
        try:
            exam_id = int(exam_id)
            if not exams.filter(id=exam_id).exists():
                exam_id = None
        except ValueError:
            exam_id = None

    # Filter students by class if class_id is provided
    if class_id:
        students = students.filter(assigned_class_id=class_id)
        subject_ids = Teacher_Subject_Class_Relation.objects.filter(assigned_classes__id=class_id).values_list('subject_id', flat=True).distinct()
        subjects = subjects.filter(id__in=subject_ids)

    # Initialize exam schedule
    schedules = ExamSchedule.objects.none()
    if exam_id:
        schedules = ExamSchedule.objects.filter(exam_id=exam_id)
        if class_id:
            schedules = schedules.filter(school_class_id=class_id)

    # Map subjects to total marks for the exam
    total_marks_dict = {sch.subject.id: sch.total_marks for sch in schedules}
    for subj in subjects:
        subj.total_marks = total_marks_dict.get(subj.id, 0)

    # Prepare marks and remarks data
    marks_data = {}
    if exam_id:
        marks_qs = Marks.objects.filter(exam_id=exam_id, student__in=students)
        remarks_qs = StudentSubjectRemark.objects.filter(exam_id=exam_id, student__in=students)

        # Build attendance lookup: {(student_id, date): status}
        attendance_qs = StudentAttendance.objects.filter(
            student__in=students,
            status__in=['Absent', 'Excused']
        )
        attendance_lookup = {
            (att.student_id, att.date): att.status
            for att in attendance_qs
        }

        # Map subject_id to exam date
        subject_exam_dates = {
            sch.subject_id: sch.date for sch in schedules
        }

        # Process marks and add attendance status to remarks
        for m in marks_qs:
            student_id = str(m.student_id)
            subject_id = str(m.subject_id)
            exam_date = subject_exam_dates.get(m.subject_id)

            display_value = m.marks_obtained  # Default mark

            # If exam date exists, check attendance status
            if exam_date:
                attendance_status = attendance_lookup.get((m.student_id, exam_date))
                if attendance_status == 'Absent':
                    display_value = 'A'
                    # Also add the attendance status to the remark column
                    marks_data.setdefault(student_id, {}).setdefault(subject_id, {})['remark'] = 'Absent'
                elif attendance_status == 'Excused':
                    display_value = 'Excused'
                    marks_data.setdefault(student_id, {}).setdefault(subject_id, {})['remark'] = 'Excused'
            
            # Store marks data
            marks_data.setdefault(student_id, {}).setdefault(subject_id, {})['marks'] = display_value

        # Store remarks data (remarks from the `StudentSubjectRemark` model)
        for r in remarks_qs:
            student_id = str(r.student_id)
            subject_id = str(r.subject_id)
            marks_data.setdefault(student_id, {}).setdefault(subject_id, {})['remark'] = r.remark

        # Ensure absentees without mark entries are marked in marks_data
        for student in students:
            for subject in subjects:
                exam_date = subject_exam_dates.get(subject.id)
                if exam_date:
                    attendance_status = attendance_lookup.get((student.id, exam_date))
                    if attendance_status == 'Absent':
                        student_id = str(student.id)
                        subject_id = str(subject.id)
                        # If no marks entry, we still show as "Absent"
                        if student_id not in marks_data or subject_id not in marks_data[student_id]:
                            marks_data.setdefault(student_id, {}).setdefault(subject_id, {})['marks'] = 'A'
                        marks_data[student_id][subject_id]['remark'] = 'Absent'
                    elif attendance_status == 'Excused':
                        student_id = str(student.id)
                        subject_id = str(subject.id)
                        # If no marks entry, we still show as "Excused"
                        if student_id not in marks_data or subject_id not in marks_data[student_id]:
                            marks_data.setdefault(student_id, {}).setdefault(subject_id, {})['marks'] = 'Excused'
                        marks_data[student_id][subject_id]['remark'] = 'Excused'

    # Prepare the context for rendering
    context = {
        'classes': classes,
        'exams': exams,
        'students': students,
        'subjects': subjects,
        'selected_class_id': class_id,
        'selected_exam_id': exam_id,
        'marks_data': json.dumps(marks_data),
    }

    return render(request, 'school_management/teacher_dashboard.html', context)

# Assuming that marks and remarks data are sent as JSON in the request body
import logging
import json
from django.db import transaction
from django.http import JsonResponse
from .models import Marks, StudentSubjectRemark, Student, Subject, Exam

# Set up logging
logger = logging.getLogger(__name__)

# Helper function to get valid objects for student, subject, and exam
def get_valid_objects(student_id, subject_id, exam_id):
    try:
        student = Student.objects.get(id=student_id)
        subject = Subject.objects.get(id=subject_id)
        exam = Exam.objects.get(id=exam_id)
        return student, subject, exam
    except Student.DoesNotExist:
        raise ValueError(f"Student with ID {student_id} not found.")
    except Subject.DoesNotExist:
        raise ValueError(f"Subject with ID {subject_id} not found.")
    except Exam.DoesNotExist:
        raise ValueError(f"Exam with ID {exam_id} not found.")

logger = logging.getLogger(__name__)
import json
import logging

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction, DatabaseError, IntegrityError

from .models import Marks, StudentSubjectRemark


logger = logging.getLogger(__name__)

# Helper functions to streamline validation
# Helper functions to streamline validation
def validate_marks_data(mark):
    if not isinstance(mark, dict):
        raise ValueError("Invalid marks data type.")

    student_id = mark.get('student_id')
    subject_id = mark.get('subject_id')
    exam_id = mark.get('exam_id')
    obtained = mark.get('marks')

    if not student_id or not subject_id or not exam_id or obtained is None:
        raise ValueError('Missing required fields in marks data.')

    if obtained == '':
        raise ValueError('Marks cannot be empty.')

    try:
        obtained = float(obtained)
    except ValueError:
        raise ValueError('Invalid marks value.')

    return {
        'student_id': student_id,
        'subject_id': subject_id,
        'exam_id': exam_id,
        'marks': obtained
    }

def validate_remarks_data(remark):
    if not isinstance(remark, dict):
        raise ValueError("Invalid remarks data type.")

    student_id = remark.get('student_id')
    subject_id = remark.get('subject_id')
    exam_id = remark.get('exam_id')
    text = remark.get('remark') or remark.get('text') or ''

    if not student_id or not subject_id or not exam_id:
        raise ValueError('Missing required fields in remarks data.')

    return {
        'student_id': student_id,
        'subject_id': subject_id,
        'exam_id': exam_id,
        'remark': str(text).strip()
    }



from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.db import transaction, DatabaseError, IntegrityError
import logging
import json

logger = logging.getLogger(__name__)

from django.db.models import Q

from django.views.decorators.http import require_POST
from django.db import transaction
from django.http import JsonResponse
from django.db.models import Q
import json
import logging

logger = logging.getLogger(__name__)

@require_POST
def update_marks(request):
    if request.content_type != 'application/json':
        return JsonResponse({'error': 'Unsupported Media Type. application/json required.'}, status=415)

    try:
        payload = json.loads(request.body)
        marks_list = payload.get('marks_data', [])
        remarks_list = payload.get('remarks_data', [])
    except Exception as e:
        logger.exception("Error parsing JSON: %s", e)
        return JsonResponse({'error': 'Invalid JSON data.'}, status=400)

    if not isinstance(marks_list, list) or not isinstance(remarks_list, list):
        return JsonResponse({'error': 'marks_data and remarks_data must be lists.'}, status=400)

    validated_marks = []
    for idx, entry in enumerate(marks_list):
        try:
            validated = validate_marks_data(entry)
            if validated:
                validated_marks.append(validated)
        except Exception as e:
            logger.warning("Skipping bad mark entry at index %d: %s", idx, e)

    validated_remarks = []
    for idx, entry in enumerate(remarks_list):
        try:
            validated = validate_remarks_data(entry)
            if validated:
                validated_remarks.append(validated)
        except Exception as e:
            logger.warning("Skipping bad remark entry at index %d: %s", idx, e)

    try:
        with transaction.atomic():
            # --- BULK UPDATE / CREATE Marks ---
            if validated_marks:
                mark_objs_to_update = []
                mark_objs_to_create = []

                existing_marks = Marks.objects.filter(
                    Q(student_id__in=[m['student_id'] for m in validated_marks]),
                    Q(subject_id__in=[m['subject_id'] for m in validated_marks]),
                    Q(exam_id__in=[m['exam_id'] for m in validated_marks])
                ).values('id', 'student_id', 'subject_id', 'exam_id')

                existing_map = {(m['student_id'], m['subject_id'], m['exam_id']): m['id'] for m in existing_marks}

                for mark in validated_marks:
                    obj_id = existing_map.get((mark['student_id'], mark['subject_id'], mark['exam_id']))
                    if obj_id:
                        mark_objs_to_update.append(Marks(
                            id=obj_id,
                            student_id=mark['student_id'],
                            subject_id=mark['subject_id'],
                            exam_id=mark['exam_id'],
                            marks_obtained=mark['marks']
                        ))
                    else:
                        mark_objs_to_create.append(Marks(
                            student_id=mark['student_id'],
                            subject_id=mark['subject_id'],
                            exam_id=mark['exam_id'],
                            marks_obtained=mark['marks']
                        ))

                if mark_objs_to_update:
                    Marks.objects.bulk_update(mark_objs_to_update, ['marks_obtained'], batch_size=100)
                if mark_objs_to_create:
                    Marks.objects.bulk_create(mark_objs_to_create, batch_size=100)

            # --- BULK UPDATE / CREATE Remarks ---
            if validated_remarks:
                remark_objs_to_update = []
                remark_objs_to_create = []

                existing_remarks = StudentSubjectRemark.objects.filter(
                    Q(student_id__in=[r['student_id'] for r in validated_remarks]),
                    Q(subject_id__in=[r['subject_id'] for r in validated_remarks]),
                    Q(exam_id__in=[r['exam_id'] for r in validated_remarks])
                ).values('id', 'student_id', 'subject_id', 'exam_id')

                existing_remark_map = {(r['student_id'], r['subject_id'], r['exam_id']): r['id'] for r in existing_remarks}

                for remark in validated_remarks:
                    obj_id = existing_remark_map.get((remark['student_id'], remark['subject_id'], remark['exam_id']))
                    if obj_id:
                        remark_objs_to_update.append(StudentSubjectRemark(
                            id=obj_id,
                            student_id=remark['student_id'],
                            subject_id=remark['subject_id'],
                            exam_id=remark['exam_id'],
                            remark=remark['remark']
                        ))
                    else:
                        remark_objs_to_create.append(StudentSubjectRemark(
                            student_id=remark['student_id'],
                            subject_id=remark['subject_id'],
                            exam_id=remark['exam_id'],
                            remark=remark['remark']
                        ))

                if remark_objs_to_update:
                    StudentSubjectRemark.objects.bulk_update(remark_objs_to_update, ['remark'], batch_size=100)
                if remark_objs_to_create:
                    StudentSubjectRemark.objects.bulk_create(remark_objs_to_create, batch_size=100)

        # 🔥 SUCCESS RESPONSE with proper message
        return JsonResponse({'success': True, 'message': 'Marks and remarks saved successfully!'})

    except Exception as e:
        logger.exception("Database error: %s", e)
        return JsonResponse({'error': 'Database error occurred.'}, status=500)

from django.shortcuts import render
from .models import Class, Subject, Exam, Teacher_Subject_Class_Relation
def exam_department(request):
    # Get the selected exam ID from the GET parameters
    selected_exam_id = request.GET.get("exam_id")
    selected_exam = None
    total_marks_dict = {}
    
    # If a valid exam id is provided, fetch the exam and its schedule.
    if selected_exam_id and selected_exam_id.isdigit():
        selected_exam = Exam.objects.filter(id=selected_exam_id).first()
        if selected_exam:
            schedules = ExamSchedule.objects.filter(exam=selected_exam)
            for schedule in schedules:
                # For each subject in this exam, store its total marks.
                total_marks_dict[schedule.subject.id] = schedule.total_marks

    # Fetch all subjects and exams.
    subjects = Subject.objects.all()
    exams = Exam.objects.all()

    # Fetch all classes and attach each class its assigned subjects.
    classes = Class.objects.all()
    for school_class in classes:
        assigned = Teacher_Subject_Class_Relation.objects.filter(assigned_classes=school_class).values_list("subject_id", flat=True)
        school_class.assigned_subjects = set(assigned)

    # Convert total_marks_dict into a list of pairs.
    total_marks_list = total_marks_dict.items()

    context = {
         "exams": exams,
         "subjects": subjects,
         "classes": classes,
         "selected_exam": selected_exam,
         "total_marks_list": total_marks_list,
    }
    return render(request, "school_management/Exam_department/exam_department.html", context)


from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.utils.dateparse import parse_date
import json
from .models import Exam, ExamSchedule, Subject, Class
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date
import json
from .models import Exam, ExamSchedule


@csrf_exempt
def create_exam(request):
    """
    Create or update an exam schedule.
    Returns error messages with reasons if any row fails.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    try:
        data = json.loads(request.body)
        exam_id = data.get("exam_id")
        exam_name = data.get("name", "").strip()
        schedule = data.get("schedule", [])

        if not schedule:
            return JsonResponse({"error": "At least one subject schedule is required!"}, status=400)

        # Get or create exam
        if exam_id:
            try:
                exam = Exam.objects.get(id=exam_id)
            except Exam.DoesNotExist:
                return JsonResponse({"error": "Selected exam does not exist!"}, status=404)
        else:
            if not exam_name:
                return JsonResponse({"error": "Exam name is required!"}, status=400)
            exam, _ = Exam.objects.get_or_create(name=exam_name)

        errors = []

        for index, entry in enumerate(schedule):
            class_id = entry.get("class_id")
            subject_id = entry.get("subject_id")
            raw_date = entry.get("date")
            raw_marks = entry.get("total_marks")

            row_id = f"[Class ID: {class_id}, Subject ID: {subject_id}]"

            # Validate date
            date = parse_date(raw_date) if raw_date else None
            if not date:
                errors.append(f"{row_id} ➜ Invalid or missing exam date.")
                continue

            # Validate marks
            try:
                total_marks = int(raw_marks)
            except (ValueError, TypeError):
                errors.append(f"{row_id} ➜ Total marks missing or not a valid number.")
                continue

            # Delete and recreate schedule
            ExamSchedule.objects.filter(
                exam=exam,
                school_class_id=class_id,
                subject_id=subject_id
            ).delete()

            ExamSchedule.objects.create(
                exam=exam,
                school_class_id=class_id,
                subject_id=subject_id,
                date=date,
                total_marks=total_marks
            )

        if errors:
            return JsonResponse({
                "error": "Some schedule entries could not be saved.",
                "details": errors
            }, status=400)

        return JsonResponse({"message": f"Exam '{exam.name}' saved successfully!"})

    except Exception as e:
        return JsonResponse({"error": "An error occurred.", "details": str(e)}, status=500)

@csrf_exempt
def delete_exam(request, exam_id):
    """Delete an exam along with its dates"""
    if request.method == "DELETE":
        exam = get_object_or_404(Exam, id=exam_id)
        exam.delete()
        return JsonResponse({"message": f"Exam '{exam.name}' deleted successfully!"})
import pandas as pd
from django.http import HttpResponse
from .models import ExamSchedule

def export_exam_schedule(request):
    """Export Exam Schedule to Excel."""
    schedules = ExamSchedule.objects.all().select_related('exam', 'school_class', 'subject')

    data = []
    for schedule in schedules:
        data.append({
            "Exam Name": schedule.exam.name,
            "Class": schedule.school_class.name,
            "Subject": schedule.subject.name,
            "Date": schedule.date.strftime("%Y-%m-%d"),
            "Total Marks": schedule.total_marks,
        })

    df = pd.DataFrame(data)

    # Create an Excel file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Exam_Schedule.xlsx"
    df.to_excel(response, index=False)

    return response
from django.http import JsonResponse
from .models import ExamSchedule

def get_exam_schedule(request, exam_id):
    """Fetch scheduled dates and marks for a selected exam"""
    schedules = ExamSchedule.objects.filter(exam_id=exam_id).select_related("school_class", "subject")

    data = [
        {
            "class_id": schedule.school_class.id,
            "subject_id": schedule.subject.id,
            "date": schedule.date.strftime("%Y-%m-%d"),
            "total_marks": schedule.total_marks,
        }
        for schedule in schedules
    ]
    
    return JsonResponse(data, safe=False)
from django.shortcuts import get_object_or_404
from django.http import JsonResponse

from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule

from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule

from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule, Class, Subject, Teacher_Subject_Class_Relation

from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule

from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule

from django.utils.dateparse import parse_date

from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule

from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import Exam, ExamSchedule

from django.db.models import OuterRef, Subquery, Q
from .models import Teacher_Subject_Class_Relation

from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation
from datetime import datetime

from datetime import datetime
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation
from datetime import datetime

import re
from datetime import datetime

def get_class_number(class_obj):
    # Extract number from class name like "10th", "9th", etc.
    match = re.search(r'\d+', class_obj.name)
    return int(match.group()) if match else 0

from django.db.models import Count
from datetime import datetime
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation

from datetime import datetime

from datetime import datetime
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation ,Student # Adjust import paths
from django.db.models.signals import post_save, post_delete
from datetime import datetime
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation,Student  # Adjust imports as per your app
from datetime import datetime
from django.db.models import Count
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation, Student, Class
from datetime import datetime
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation
from datetime import datetime
from django.shortcuts import render

import re
from datetime import datetime
from django.shortcuts import render

import re
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation


from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render
from .models import Exam, ExamSchedule, Teacher_Subject_Class_Relation  # Adjust imports as needed

def exams_by_date_view(request):
    def parse_date_safe(date_str):
        cleaned_str = date_str.replace("Sept.", "Sep").replace("Sept", "Sep")
        for fmt in ("%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"):
            try:
                return datetime.strptime(cleaned_str, fmt).date()
            except ValueError:
                continue
        return None

    def extract_class_number(class_name):
        match = re.match(r'(\d+)', class_name)
        if match:
            return int(match.group(1))
        return 9999

    def export_schedules_to_excel(schedules):
        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Schedule"

        headers = [
            "Class",
            "Number of Students",
            "Subject",
            "Total Marks",
            "Invigilator",
            "Invigilator Signature",
            "Subject Teacher",
            "Teacher Received Signature",
        ]
        ws.append(headers)

        col_widths = [15, 20, 20, 15, 25, 25, 25, 25]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for s in schedules:
            row = [
                s['school_class'].name,
                s['num_students'],
                s['subject'].name,
                s['total_marks'],
                s.get('invigilator_name', ''),
                "",
                s.get('subject_teacher_name', ''),
                "",
            ]
            ws.append(row)

        return wb

    exams = Exam.objects.all()
    selected_exam_id = request.GET.get('exam_id')
    selected_date = request.GET.get('date')
    selected_subject_id = request.GET.get('subject_id')
    export_format = request.GET.get('export')  # e.g. 'excel'
    schedules = []
    exam_dates = []
    subjects = []

    if selected_exam_id:
        exam_dates_qs = ExamSchedule.objects.filter(exam_id=selected_exam_id) \
                                            .values_list('date', flat=True) \
                                            .distinct()
        exam_dates = sorted(set(exam_dates_qs))

        subjects_qs = ExamSchedule.objects.filter(exam_id=selected_exam_id) \
                                         .values('subject__id', 'subject__name') \
                                         .distinct()
        subjects = [{'id': s['subject__id'], 'name': s['subject__name']} for s in subjects_qs]

        if selected_date:
            parsed_date = parse_date_safe(selected_date)

            if parsed_date:
                schedule_qs = ExamSchedule.objects.filter(
                    exam_id=selected_exam_id,
                    date=parsed_date
                ).select_related('school_class', 'subject')

                if selected_subject_id:
                    schedule_qs = schedule_qs.filter(subject_id=selected_subject_id)

                for entry in schedule_qs:
                    subject_teacher_relation = Teacher_Subject_Class_Relation.objects.filter(
                        subject=entry.subject,
                        assigned_classes=entry.school_class,
                        teacher__isnull=False
                    ).select_related('teacher').first()

                    schedules.append({
                        "school_class": entry.school_class,
                        "num_students": entry.school_class.students.count(),
                        "subject": entry.subject,
                        "total_marks": entry.total_marks,
                        "subject_teacher_name": subject_teacher_relation.teacher.name if subject_teacher_relation else "",
                        "invigilator_name": getattr(entry, 'invigilator_name', ''),
                    })

                schedules.sort(key=lambda s: extract_class_number(s['school_class'].name))

    # Calculate total number of students
    total_students = sum(s['num_students'] for s in schedules) if schedules else 0

    if export_format == 'excel' and schedules:
        wb = export_schedules_to_excel(schedules)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Exam_Schedule_{selected_exam_id}_{selected_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    context = {
        "exams": exams,
        "selected_exam_id": selected_exam_id,
        "selected_date": selected_date,
        "exam_dates": exam_dates,
        "subjects": subjects,
        "selected_subject_id": selected_subject_id,
        "schedules": schedules,
        "total_students": total_students,  # Pass total students to template
    }
    return render(request, "school_management/Exam_department/exams_by_date.html", context)

def get_student_profile(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    data = {
        "name": student.name,
        "roll_number": student.roll_number,
        "class": student.assigned_class.name,
        "phone": student.phone,
        "address": student.address,
    }

    return JsonResponse(data)
def get_student_marks(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    marks = Marks.objects.filter(student=student)

    data = [
        {
            "subject": mark.subject.name,
            "exam": mark.exam.name,
            "marks_obtained": mark.marks_obtained,
            "total_marks": mark.total_obtained_marks,
        }
        for mark in marks
    ]

    return JsonResponse(data, safe=False)
def get_student_timetable(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    timetable = Timetable.objects.filter(class_obj=student.assigned_class)

    data = [
        {
            "day": entry.day,
            "period": entry.period,
            "subject": entry.teacher_subject.subject.name if entry.teacher_subject else "Free",
            "teacher": entry.teacher_subject.teacher.name if entry.teacher_subject else "None",
        }
        for entry in timetable
    ]

    return JsonResponse(data, safe=False)
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import  Student, Marks, Timetable

from django.contrib.auth import authenticate, login
from django.contrib import messages
from .models import Parent, Teacher, Student  # Import required models
def team_login_pg(request):
    return render(request, "school_management/team_login.html")




from django.contrib.auth.hashers import check_password, make_password
from django.utils.timezone import now
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Parent, Teacher


def team_login(request):
    if request.method == "POST":
        login_id = request.POST.get("login_id")
        password = request.POST.get("password")

        print(f"🔍 DEBUG: Received Login ID: {login_id}")  

        # ✅ Fix: Use "username" instead of "login_id"
        user = Parent.objects.filter(username=login_id).first()
        if user:
            print(f"🔍 DEBUG: Found Parent - {user.username}")
            print(f"🔍 Stored Hashed Password: {user.password}")

            if check_password(password, user.password):  # Verify hashed password
                print("✅ DEBUG: Password Matched!")
                request.session["user_id"] = user.id  
                request.session["user_role"] = "parent_dashboard"  
                
                user.last_login = now()
                user.save(update_fields=["last_login"])

                return redirect("parent_dashboard")  

        # Check if user is a Teacher
        user = Teacher.objects.filter(username=login_id).first()
        if user:
            print(f"🔍 DEBUG: Found Teacher - {user.username}")

            if check_password(password, user.password):  
                print("✅ DEBUG: Password Matched!")
                request.session["user_id"] = user.id  
                request.session["user_role"] = "teacher_dashboard"
                
                user.last_login = now()
                user.save(update_fields=["last_login"])

                return redirect("teacher_dashboard")

        print("❌ DEBUG: Invalid Login ID or Password")
        messages.error(request, "Invalid Login ID or Password!")
        return redirect("team_login")

    return render(request, "school_management/team_login.html")


from django.contrib.auth import get_user_model
from django.contrib.auth.backends import ModelBackend
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .models import  Teacher

User = get_user_model()
from django.shortcuts import render, redirect
from .models import Parent, Student

from django.shortcuts import render, redirect
from .models import Parent, Student
from django.shortcuts import render, get_object_or_404, redirect
from .models import Parent, Student, Subject, Exam, ExamSchedule, Marks

from django.shortcuts import render, redirect
from .models import Parent, Student, Subject, Exam, ExamSchedule, Marks
def parent_dashboard(request):
    parent_id = request.session.get("user_id")
    user_role = request.session.get("user_role")

    if not parent_id or user_role != "parent_dashboard":
        return redirect("team_login")

    try:
        parent = Parent.objects.get(id=parent_id)
        children = Student.objects.filter(parent=parent)

        student_data = []
        for student in children:
            subjects = Subject.objects.filter(
                id__in=ExamSchedule.objects.filter(school_class=student.assigned_class)
                .values_list("subject_id", flat=True)
            )

            exam_types = {
                exam.name: exam.schedules.first().total_marks if exam.schedules.exists() else "-"
                for exam in Exam.objects.all()
            }

            student_marks = []
            for subject in subjects:
                marks_entry = {"subject_name": subject.name}
                for exam_name in exam_types.keys():
                    obtained_marks = Marks.objects.filter(
                        student=student, subject=subject, exam__name=exam_name
                    ).values_list("marks_obtained", flat=True).first() or "N/A"
                    marks_entry[exam_name] = obtained_marks
                student_marks.append(marks_entry)  # ✅ Make it a list of dicts

            student_data.append({
                "student": student,
                "exam_types": exam_types,
                "student_marks": student_marks,
            })

        return render(request, "school_management/parent_dashboard.html", {
            "parent": parent,
            "student_data": student_data
        })

    except Parent.DoesNotExist:
        return redirect("team_login")



def parent_logout(request):
    """Logs out the parent and clears the session"""
    request.session.flush()  # Clears all session data

    return redirect("team_login")  # Redirect back to login page




from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from .models import Exam, Subject, Student, InternalAssessment, Class

from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from .models import Exam, Subject, Student, InternalAssessment, Class

@csrf_exempt
def assessment_table(request):
    print(">>> assessment_table view triggered")

    # —–– Load all exams for dropdown
    exams = Exam.objects.all()
    exam_id = request.GET.get('exam_id')
    current_exam = None
    if exam_id:
        try:
            current_exam = exams.get(pk=exam_id)
        except Exam.DoesNotExist:
            messages.error(request, "Selected exam not found.")
            return redirect(request.path)

    # —–– Load subjects & subsubjects
    subjects = Subject.objects.prefetch_related("sub_subjects").all()

    # —–– Class filter
    class_filter = request.GET.get('class_filter')
    students = Student.objects.filter(assigned_class__id=class_filter) if class_filter else Student.objects.none()

    # —–– Only proceed if both exam and class are chosen
    assessments = InternalAssessment.objects.none()
    if current_exam and class_filter:
        assessments = InternalAssessment.objects.filter(
            exam=current_exam,
            school_class__id=class_filter
        ).select_related('student','subject','subsubject')

    # —–– Build lookup for max_marks
    max_marks_lookup = {
        f"{a.subject.id}_{(a.subsubject.id if a.subsubject else 'none')}": a.max_marks
        for a in assessments
    }

    # —–– Build subject_groups
    subject_groups = []
    for subject in subjects:
        raw = list(subject.sub_subjects.all()) or [None]
        subs_info = []
        for idx, sub in enumerate(raw):
            key = f"{subject.id}_{(sub.id if sub else 'none')}"
            subs_info.append({
                'obj': sub,
                'index': idx,
                'subject_id': subject.id,
                'subsubject_id': sub.id if sub else 'none',
                'max_marks': max_marks_lookup.get(key)
            })
        subject_groups.append({'subject': subject, 'subs_info': subs_info})

    # —–– Handle POST (save marks)
    if request.method == "POST" and current_exam and class_filter:
        updated = 0
        for student in students:
            for group in subject_groups:
                for si in group['subs_info']:
                    field = f"marks_{student.id}_{si['subject_id']}_{si['subsubject_id']}"
                    val = request.POST.get(field, "").strip()
                    if val:
                        obj, created = InternalAssessment.objects.get_or_create(
                            student=student,
                            subject=group['subject'],
                            subsubject=si['obj'],
                            school_class=student.assigned_class,
                            exam=current_exam,
                            defaults={'obtained_marks': val, 'max_marks': si['max_marks'] or 0}
                        )
                        if not created:
                            obj.obtained_marks = val
                            obj.save()
                        updated += 1
        if updated:
            messages.success(request, f"{updated} mark{'s' if updated>1 else ''} saved for {current_exam.name}.")
        else:
            messages.info(request, "No marks entered to save.")
        # redirect preserving filters
        return redirect(f"{request.path}?exam_id={exam_id}&class_filter={class_filter}")

    # —–– Build table_data
    table_data = []
    for student in students:
        row = {'student': student, 'cells': []}
        for group in subject_groups:
            cells = []
            for si in group['subs_info']:
                a = assessments.filter(
                    student=student,
                    subject=group['subject'],
                    subsubject=si['obj']
                ).first()
                cells.append({
                    'value': a.obtained_marks if a else '',
                    'subject_id': si['subject_id'],
                    'subsubject_id': si['subsubject_id']
                })
            row['cells'].append(cells)
        table_data.append(row)

    return render(request, "school_management/Tr_student_assement_entry.html", {
        'exams': exams,
        'current_exam': current_exam,
        'school_classes': Class.objects.all(),
        'class_filter': class_filter,
        'current_year': 2025,
        'students': students,
        'subject_groups': subject_groups,
        'table_data': table_data,
    })


def internal_assessment_pg(request):
    return render(request, 'school_management/Exam_dep_internal_assment.html')

from django.shortcuts import render
from .models import Class, ClassSubSubjectAssignment

def debug_class_subject_assignment(request):
    class_id = request.GET.get('class_id')

    if not class_id:
        return render(request, 'debug.html', {"message": "No class ID provided."})

    try:
        school_class = Class.objects.get(id=class_id)
    except Class.DoesNotExist:
        return render(request, 'debug.html', {"message": "Class not found."})

    assignments = ClassSubSubjectAssignment.objects.filter(school_class=school_class)

    return render(request, 'debug.html', {
        "school_class": school_class,
        "assignments": assignments,
        "count": assignments.count()
    })


from django.shortcuts import render, redirect
from django.utils.timezone import now
from .models import Class, Student, StudentAttendance




from django.shortcuts import render
from .models import Class, Student, StudentAttendance
from django.utils.timezone import now
def attendance_view(request):
    date = now().date()
    selected_class_id = request.GET.get('class_id')
    classes = Class.objects.all()
    students = []
    selected_class_name = None

    if selected_class_id:
        try:
            school_class = Class.objects.get(id=selected_class_id)
            students = list(Student.objects.filter(assigned_class=school_class).order_by('roll_number'))
            attendance_records = StudentAttendance.objects.filter(school_class=school_class, date=date)
            
            # Create a mapping of student ID to attendance status
            attendance_status_map = {
                record.student.id: record.status for record in attendance_records
            }

            # Assign the attendance status to each student
            for student in students:
                student.attendance_status = attendance_status_map.get(student.id, '')  # Empty if not marked

            selected_class_name = school_class.name
        except Class.DoesNotExist:
            messages.error(request, "Class not found.")
            return redirect('attendance_view')

    return render(request, 'school_management/students_attendance.html', {
        'classes': classes,
        'selected_class_id': selected_class_id,
        'selected_class_name': selected_class_name,
        'students': students,
        'date': date,
    })
def save_attendance(request):
    if request.method == 'POST':
        selected_class_id = request.POST.get('class_id')
        school_class = Class.objects.get(id=selected_class_id)
        students = Student.objects.filter(assigned_class=school_class)
        date = now().date()

        try:
            for student in students:
                status = request.POST.get(f'status_{student.id}')

                # Ensure the status is set
                if status is None:
                    # Set a default status (optional) or skip saving this record
                    raise ValueError(f"Status for student {student.name} is not selected.")

                StudentAttendance.objects.update_or_create(
                    student=student,
                    school_class=school_class,
                    date=date,
                    defaults={'status': status}
                )

            messages.success(request, "Attendance saved successfully.")
        except Exception as e:
            messages.error(request, f"Failed to save attendance: {str(e)}")

        return redirect('attendance_view')


from django.shortcuts import render
from .models import Class, Student
from datetime import datetime

from django.shortcuts import render
from .models import Class, StudentAttendance
from datetime import datetime

def attendance_list_view(request):
    classes = Class.objects.all()
    selected_class_id = request.GET.get('class_id')
    selected_date = request.GET.get('date')
    attendance_records = []
    selected_class_name = ""

    if selected_class_id and selected_date:
        try:
            selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
            attendance_records = StudentAttendance.objects.filter(
                school_class_id=selected_class_id,
                date=selected_date_obj
            ).select_related('student')
            selected_class_name = Class.objects.get(id=selected_class_id).name
        except Exception:
            pass

    return render(request, 'school_management/attendance_list.html', {
        'classes': classes,
        'selected_class_id': selected_class_id,
        'selected_date': selected_date,
        'selected_class_name': selected_class_name,
        'attendance_records': attendance_records,
    })

from django.shortcuts import render
from django.http import Http404
from .models import InternalAssessment, Marks, StudentAttendance
from django.db.models import F




from django.db.models import Sum, F, FloatField, ExpressionWrapper
from .models import Marks, Student

def student_assessment_report(request):
    class_id = request.GET.get('class_id')
    search_name = request.GET.get('search_name', '')

    students = Student.objects.all()
    if class_id:
        students = students.filter(assigned_class__id=class_id)
    if search_name:
        students = students.filter(name__icontains=search_name)

    assessments = []

    for student in students:
        student_marks = Marks.objects.filter(student=student)

        total_obtained = 0
        total_possible = 0

        for mark in student_marks:
            total_obtained += mark.marks_obtained or 0  # handle None

            # Get the corresponding exam schedule for this subject and class
            try:
                schedule = ExamSchedule.objects.get(
                    exam=mark.exam,
                    subject=mark.subject,
                    school_class=student.assigned_class
                )
                total_possible += schedule.total_marks
            except ExamSchedule.DoesNotExist:
                continue  # Skip if no schedule found

        if total_possible > 0:
            average_percentage = (total_obtained / total_possible) * 100
        else:
            average_percentage = 0

        assessments.append({
            'student': student,
            'average_percentage': round(average_percentage, 2),
        })

    context = {
        'assessments': assessments,
        'classes': Class.objects.all(),
        'selected_class': class_id,
        'search_name': search_name,
    }

    return render(request, 'school_management/students/student_assessment_ui.html', context)

from django.shortcuts import get_object_or_404, render
from .models import Student, StudentAttendance, Exam, Marks, ExamSchedule, StudentSubjectRemark
from django.db.models import Count
from django.shortcuts import render, get_object_or_404
from .models import Student, StudentAttendance, Exam, Marks, ExamSchedule, StudentSubjectRemark
from django.db.models import Count
from django.shortcuts import get_object_or_404, render
from django.shortcuts import get_object_or_404, render
from django.db.models import Count
from datetime import date
from django.shortcuts import render, get_object_or_404
from django.db.models import Count, Q
from .models import Student, StudentAttendance, Exam, Marks, ExamSchedule, StudentSubjectRemark

from django.shortcuts import render, get_object_or_404
from django.db.models import Count, Q
from .models import Student, StudentAttendance, Exam, Marks, ExamSchedule, StudentSubjectRemark

from django.shortcuts import render, get_object_or_404
from django.db.models import Count, Q
from .models import Student, StudentAttendance, Marks, Exam, ExamSchedule, StudentSubjectRemark, AcademicYear
from django.shortcuts import render, get_object_or_404
from django.db.models import Count, Q
from datetime import date
from .models import (
    Student, StudentAttendance, Exam, Marks,
    ExamSchedule, StudentSubjectRemark, AcademicYear
)

def student_detail(request, student_id):
    # Fetch student with related parent, class, and academic year
    student = get_object_or_404(
        Student.objects.select_related('parent', 'assigned_class', 'academic_year'),
        id=student_id
    )
    
    parent = student.parent
    class_name = student.assigned_class.name
    academic_year = student.academic_year

    # === Attendance Statistics ===
    attendance_stats = StudentAttendance.objects.filter(student=student).aggregate(
        total_present=Count('status', filter=Q(status='Present')),
        total_absent=Count('status', filter=Q(status='Absent')),
        total_late=Count('status', filter=Q(status='Late')),
        total_excused=Count('status', filter=Q(status='Excused'))
    )

    total_present = attendance_stats.get('total_present') or 0
    total_absent = attendance_stats.get('total_absent') or 0
    total_late = attendance_stats.get('total_late') or 0
    total_excused = attendance_stats.get('total_excused') or 0

    total_working_days = academic_year.working_days_upto_today()
    attendance_percentage = round((total_present / total_working_days) * 100, 2) if total_working_days else 0

    # === Exam Selection ===
    selected_exam_id = request.GET.get('exam')
    exams = Exam.objects.all().order_by('-id')  # For exam dropdown

    # === Marks and Exam Schedules ===
    marks_records = Marks.objects.filter(student=student).select_related('subject')
    if selected_exam_id:
        marks_records = marks_records.filter(exam_id=selected_exam_id)

    exam_schedules = ExamSchedule.objects.filter(
        school_class=student.assigned_class,
        exam_id=selected_exam_id
    ) if selected_exam_id else []

    schedule_dict = {s.subject.id: s for s in exam_schedules}
    
    remarks = StudentSubjectRemark.objects.filter(student=student)
    if selected_exam_id:
        remarks = remarks.filter(exam_id=selected_exam_id)
    remark_dict = {r.subject.id: r.remark for r in remarks}

    # === Prepare Marks Data ===
    marks_data = {}
    total_marks_obtained = 0
    total_max_marks = 0

    for mark in marks_records:
        subject_id = mark.subject.id
        schedule = schedule_dict.get(subject_id)
        total_marks = schedule.total_marks if schedule else 0
        obtained = mark.marks_obtained
        percentage = round((obtained / total_marks) * 100, 2) if total_marks > 0 else 0

        marks_data[subject_id] = {
            'subject': mark.subject,
            'marks_obtained': obtained,
            'total_marks': total_marks,
            'percentage': percentage,
            'remark': remark_dict.get(subject_id, 'No remark available')
        }

        total_marks_obtained += obtained
        total_max_marks += total_marks

    overall_percentage = round((total_marks_obtained / total_max_marks) * 100, 2) if total_max_marks > 0 else 0

    # === Context for Template ===
    context = {
        'student': student,
        'parent': parent,
        'class_name': class_name,
        'attendance_counts': {
            'Present': total_present,
            'Absent': total_absent,
            'Late': total_late,
            'Excused': total_excused,
        },
        'attendance_percentage': attendance_percentage,
        'total_working_days': total_working_days,
        'exams': exams,
        'selected_exam_id': selected_exam_id,
        'marks_data': marks_data,
        'exam_schedules': exam_schedules,
        'total_marks_obtained': total_marks_obtained,
        'total_max_marks': total_max_marks,
        'percentage': overall_percentage,
    }

    return render(request, 'school_management/students/student_detail.html', context)
from django.shortcuts import render, get_object_or_404
from .models import Student, Exam, Marks, ExamSchedule, StudentSubjectRemark
from django.shortcuts import render, get_object_or_404
from .models import Student, Exam, Marks, ExamSchedule, StudentSubjectRemark, SchoolProfile
from django.shortcuts import render, get_object_or_404
from .models import Student, Exam, Marks, ExamSchedule, StudentSubjectRemark, SchoolProfile
from django.shortcuts import render, get_object_or_404
from .models import (
    Student, Exam, Marks, ExamSchedule, StudentSubjectRemark,
    SchoolProfile
)

from django.shortcuts import get_object_or_404, render
from .models import Student, SchoolProfile, Exam, Marks, ExamSchedule, StudentSubjectRemark, Class

def student_marksheet_view(request, student_id):
    # Fetch student with related info (including assigned_class)
    student = get_object_or_404(
        Student.objects.select_related('parent', 'assigned_class', 'academic_year'),
        id=student_id
    )
    
    # Fetch the first (or only) school profile
    school_profile = SchoolProfile.objects.first()
    
    exam1_id = request.GET.get('exam1')
    exam2_id = request.GET.get('exam2')
    
    exams = Exam.objects.all().order_by('-id')
    
    exam1 = Exam.objects.filter(id=exam1_id).first() if exam1_id else None
    exam2 = Exam.objects.filter(id=exam2_id).first() if exam2_id else None
    
    marks_by_exam = {}
    
    for exam_id in [exam1_id, exam2_id]:
        if exam_id:
            marks_records = Marks.objects.filter(student=student, exam_id=exam_id).select_related('subject')
            schedules = ExamSchedule.objects.filter(exam_id=exam_id, school_class=student.assigned_class)
            remarks = StudentSubjectRemark.objects.filter(student=student, exam_id=exam_id)
    
            schedule_dict = {s.subject.id: s for s in schedules}
            remark_dict = {r.subject.id: r.remark for r in remarks}
    
            marks_data = []
            for mark in marks_records:
                subject = mark.subject
                total_marks = schedule_dict.get(subject.id).total_marks if schedule_dict.get(subject.id) else 0
                obtained = mark.marks_obtained
                percentage = round((obtained / total_marks) * 100, 2) if total_marks else 0
                marks_data.append({
                    'subject': subject.name,
                    'obtained': obtained,
                    'total': total_marks,
                    'percentage': percentage,
                    'remark': remark_dict.get(subject.id, ''),
                })
    
            marks_by_exam[str(exam_id)] = marks_data
    
    marks_exam1 = marks_by_exam.get(str(exam1_id), []) if exam1_id else []
    marks_exam2 = marks_by_exam.get(str(exam2_id), []) if exam2_id else []
    
    # Get the assigned class object (to get class teacher)
    assigned_class = student.assigned_class
    
    # Get class teacher name if exists
    class_teacher_name = ""
    if assigned_class and hasattr(assigned_class, 'class_teacher') and assigned_class.class_teacher:
        class_teacher_name = assigned_class.class_teacher.name
    
    context = {
        'student': student,
        'exams': exams,
        'exam1_id': exam1_id,
        'exam2_id': exam2_id,
        'exam1_name': exam1.name if exam1 else '',
        'exam2_name': exam2.name if exam2 else '',
        'marks_exam1': marks_exam1,
        'marks_exam2': marks_exam2,
        'school_name': school_profile.name if school_profile else "Your School Name",
        'school_address': school_profile.address if school_profile else "School Address Here",
        'school_phone': school_profile.phone_number if school_profile else "",
        'school_email': school_profile.email if school_profile else "",
        'principal_name': school_profile.principal_name if school_profile else "",
        'school_logo_url': school_profile.logo.url if (school_profile and school_profile.logo) else "",
        'class_teacher_name': class_teacher_name,
        'selected_class_obj': assigned_class,  # pass for template access if needed
        'now': timezone.localtime(timezone.now())
    }
    
    return render(request, 'school_management/students/student_marksheet.html', context)

from django.shortcuts import render, get_object_or_404
from .models import Student, Exam, InternalAssessment, StudentSubjectRemark
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
import json
from .models import TransferredStudent
# views.py
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Student, Parent, Marks, TransferredStudent
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
import json

def transfer_student(request, student_id):
    student = get_object_or_404(Student, pk=student_id)

    if request.method == "POST":
        marks = student.marks.all()
        marks_data = []
        for mark in marks:
            marks_data.append({
                "subject": mark.subject.name,
                "exam": mark.exam.name,
                "marks_obtained": mark.marks_obtained,
                "total_obtained_marks": mark.total_obtained_marks,
                "remarks": mark.remarks,
            })
        marks_data_json = json.dumps(marks_data)

        student_class_name = str(student.assigned_class) if student.assigned_class else None

        parent = student.parent
        parent_name = parent.name if parent else None
        parent_email = parent.email if parent else None
        parent_phone = parent.phone if parent else None

        # Update this based on actual fields in your TransferredStudent model
        transferred_student = TransferredStudent.objects.create(
            name=student.name,
            # pen=student.pen,  # Remove if no such field
            # aadhar=student.aadhar,
            # phone=student.phone,
            parent_name=parent_name,
            parent_email=parent_email,
            parent_phone=parent_phone,
            student_class=student_class_name,
            transferred_on=timezone.now(),
            marks_snapshot=marks_data_json,
        )

        marks.delete()
        student.delete()

        return redirect('transferred_student_list')

    return redirect('student_detail', student_id=student_id)


def student_internal_assessment(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    exams = Exam.objects.all()

    selected_exam_id = request.GET.get('exam')
    selected_exam = None
    grouped_subjects = {}

    if selected_exam_id:
        selected_exam = Exam.objects.get(id=selected_exam_id)

        # Fetch all internal assessments for the student in the selected exam
        assessments = InternalAssessment.objects.filter(student=student, exam=selected_exam)

        for assessment in assessments:
            subject_name = assessment.subject.name  # <-- Grouping by correct subject
            subsubject_name = assessment.subsubject.name if assessment.subsubject else "General"

            percentage = None
            if assessment.max_marks and assessment.obtained_marks is not None:
                percentage = round((assessment.obtained_marks / assessment.max_marks) * 100, 2)

            if subject_name not in grouped_subjects:
                grouped_subjects[subject_name] = []

            grouped_subjects[subject_name].append({
                "subsubject": subsubject_name,
                "marks_obtained": assessment.obtained_marks,
                "total_marks": assessment.max_marks,
                "percentage": percentage,
            })

    context = {
        "student": student,
        "exams": exams,
        "grouped_subjects": grouped_subjects,
        "selected_exam": selected_exam,
    }

    print("Grouped Subjects:", grouped_subjects)
    return render(request, "school_management/internal_assessment_report.html", context)


from .models import (
    Class,
    Student,
    Subject,
    Teacher_Subject_Class_Relation,
    SkillOrDevelopmentArea,
    SyllabusPerformance,
    SoftSkillPerformance,
)
from django.shortcuts import render, redirect
from .models import (
    Class,
    Student,
    Subject,
    Teacher_Subject_Class_Relation,
    SkillOrDevelopmentArea,
    SyllabusPerformance,
    SoftSkillPerformance,
)

from django.contrib import messages
from django.shortcuts import render, redirect

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import (
    Class, Student, Subject,
    Teacher_Subject_Class_Relation,
    SyllabusPerformance, SoftSkillPerformance,
    SkillOrDevelopmentArea
)

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import (
    Class, Student, Subject, SkillOrDevelopmentArea,
    SyllabusPerformance, SoftSkillPerformance,
    Teacher_Subject_Class_Relation
)

from datetime import date
from django.shortcuts import render, redirect
from django.contrib import messages

from datetime import date
from django.utils.dateparse import parse_date  # <-- Import this

def solution_based_sorting_view(request):
    selected_class_id = request.GET.get('class_id')
    selected_date_str = request.GET.get('date_filter')
    selected_date = parse_date(selected_date_str) if selected_date_str else None  # <-- Parse date safely

    # Handle creation of new options
    if request.method == "POST":
        if 'create_type' in request.POST:
            new_name = request.POST.get('new_value')
            create_type = request.POST.get('create_type')
            if new_name:
                if create_type == "syllabus":
                    SyllabusPerformance.objects.get_or_create(name=new_name)
                    messages.success(request, "New syllabus option added successfully!")
                elif create_type == "softskill":
                    SoftSkillPerformance.objects.get_or_create(name=new_name)
                    messages.success(request, "New soft skill option added successfully!")
            return redirect(f"{request.path}?class_id={selected_class_id or ''}&date_filter={selected_date_str or ''}")

        elif 'delete_type' in request.POST:
            delete_name = request.POST.get('delete_value')
            delete_type = request.POST.get('delete_type')
            if delete_name:
                if delete_type == "syllabus":
                    SyllabusPerformance.objects.filter(name=delete_name).delete()
                    messages.success(request, "Syllabus option deleted successfully!")
                elif delete_type == "softskill":
                    SoftSkillPerformance.objects.filter(name=delete_name).delete()
                    messages.success(request, "Soft skill option deleted successfully!")
            return redirect(f"{request.path}?class_id={selected_class_id or ''}&date_filter={selected_date_str or ''}")

        # Save posted assessments
        else:
            assessment_date_str = request.POST.get('assessment_date') or str(date.today())
            assessment_date = parse_date(assessment_date_str)

            if selected_class_id:
                students = Student.objects.filter(assigned_class_id=selected_class_id)
                subjects = Subject.objects.filter(
                    id__in=Teacher_Subject_Class_Relation.objects.filter(
                        assigned_classes__id=selected_class_id
                    ).values_list("subject_id", flat=True)
                )

                for student in students:
                    for subject in subjects:
                        prefix = f"{student.id}_{subject.id}"
                        syllabus_val = request.POST.get(f"{prefix}_syllabus", "")
                        softskill_val = request.POST.get(f"{prefix}_softskill", "")
                        SkillOrDevelopmentArea.objects.update_or_create(
                            student=student,
                            school_class_id=selected_class_id,
                            subject=subject,
                            assessment_date=assessment_date,
                            defaults={
                                'syllabus_assessment': syllabus_val,
                                'softskill_assessment': softskill_val
                            }
                        )
                messages.success(request, "Assessment data saved successfully!")
                return redirect(f"{request.path}?class_id={selected_class_id}&date_filter={assessment_date_str}")

    # Continue with GET logic
    classes = Class.objects.all()
    students = Student.objects.none()
    subjects = Subject.objects.none()
    assessment_rows = []

    syllabus_options = SyllabusPerformance.objects.all()
    softskill_options = SoftSkillPerformance.objects.all()

    if selected_class_id:
        students = Student.objects.filter(assigned_class_id=selected_class_id)
        subject_relations = Teacher_Subject_Class_Relation.objects.filter(
            assigned_classes__id=selected_class_id
        ).select_related("subject")
        subject_ids = subject_relations.values_list("subject_id", flat=True)
        subjects = Subject.objects.filter(id__in=subject_ids)

        assessment_query = SkillOrDevelopmentArea.objects.filter(
            school_class_id=selected_class_id,
            subject__in=subjects
        )
        if selected_date:
            assessment_query = assessment_query.filter(assessment_date=selected_date)

        existing_assessments = {
            (a.student_id, a.subject_id): a
            for a in assessment_query
        }

        for student in students:
            for subject in subjects:
                assessment = existing_assessments.get((student.id, subject.id))
                assessment_rows.append({
                    'student': student,
                    'subject': subject,
                    'student_id': student.id,
                    'subject_id': subject.id,
                    'syllabus': assessment.syllabus_assessment if assessment else '',
                    'softskill': assessment.softskill_assessment if assessment else '',
                    'date': assessment.assessment_date if assessment else '',
                })

    return render(request, "school_management/solution_based_entry.html", {
        'classes': classes,
        'selected_class_id': selected_class_id,
        'selected_date': selected_date_str,
        'students': students,
        'subjects': subjects,
        'assessment_rows': assessment_rows,
        'syllabus_options': syllabus_options,
        'softskill_options': softskill_options,
        'today': date.today(),
    })
# views.py

from django.shortcuts import render, get_object_or_404
from django.shortcuts import render
from .models import AcademicYear

# views.py

from django.shortcuts import render, redirect
from .models import AcademicYear
from django.contrib import messages
from datetime import datetime
from django.shortcuts import render
from django.contrib import messages
from datetime import datetime
from .models import AcademicYear

def create_academic_year(request):
    name = ''
    start_date = ''
    end_date = ''

    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if not name or not start_date or not end_date:
            messages.error(request, "All fields are required.")
        else:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()

                if end < start:
                    messages.error(request, "End date cannot be before start date.")
                else:
                    AcademicYear.objects.create(name=name, start_date=start, end_date=end)
                    messages.success(request, "Academic year created successfully.")
                    # Reset form values after success
                    name = ''
                    start_date = ''
                    end_date = ''
            except ValueError:
                messages.error(request, "Invalid date format.")

    context = {
        'name': name,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'school_management/create_academic_year.html', context)


def academic_year_list(request):
    years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'school_management/academic_year_list.html', {'years': years})

from datetime import timedelta
from django.shortcuts import render, get_object_or_404
from .models import AcademicYear, Holiday
from datetime import timedelta

def working_days_summary(request, year_id):
    academic_year = get_object_or_404(AcademicYear, id=year_id)

    start_date = academic_year.start_date
    end_date = academic_year.end_date

    total_days = (end_date - start_date).days + 1
    current_date = start_date

    # Get holidays
    holidays = academic_year.holiday_set.all().order_by('date')
    manual_holidays = set(holiday.date for holiday in holidays)

    sundays = 0
    manual_holiday_count = 0
    working_days = 0

    while current_date <= end_date:
        if current_date.weekday() == 6:  # Sunday
            sundays += 1
        elif current_date in manual_holidays:
            manual_holiday_count += 1
        else:
            working_days += 1
        current_date += timedelta(days=1)

    context = {
        'academic_year': academic_year,
        'total_days': total_days,
        'sundays': sundays,
        'manual_holiday_count': manual_holiday_count,
        'working_days': working_days,
        'holidays': holidays,  # ← important fix
    }

    return render(request, 'school_management/working_days_summary.html', context)

from django.shortcuts import render, get_object_or_404, redirect
from .models import AcademicYear
from django.contrib import messages
from datetime import datetime

from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import AcademicYear

def edit_academic_year(request, pk):
    academic_year = get_object_or_404(AcademicYear, pk=pk)

    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if not name or not start_date or not end_date:
            messages.error(request, "All fields are required.")
        else:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()

                if end < start:
                    messages.error(request, "End date cannot be before start date.")
                else:
                    academic_year.name = name
                    academic_year.start_date = start
                    academic_year.end_date = end
                    academic_year.save()
                    messages.success(request, "Academic year updated successfully.")
                    return redirect('academic_year_list')
            except ValueError:
                messages.error(request, "Invalid date format.")

    return render(request, 'school_management/edit_academic_year.html', {'academic_year': academic_year})

def add_holiday(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        date = request.POST.get('date')
        if not name or not date:
            messages.error(request, "All fields are required.")
        else:
            # You'll need to create a Holiday model first.
            from .models import Holiday
            try:
                holiday_date = datetime.strptime(date, '%Y-%m-%d').date()
                Holiday.objects.create(name=name, date=holiday_date)
                messages.success(request, "Holiday added successfully.")
                return redirect('holiday_add')
            except ValueError:
                messages.error(request, "Invalid date format.")

    return render(request, 'school_management/add_holiday.html')
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import datetime
from .models import Holiday, AcademicYear

def add_holiday(request):
    academic_years = AcademicYear.objects.all()

    if request.method == 'POST':
        reason = request.POST.get('reason')
        date = request.POST.get('date')
        academic_year_id = request.POST.get('academic_year')

        if not (reason and date and academic_year_id):
            messages.error(request, "All fields are required.")
        else:
            try:
                holiday_date = datetime.strptime(date, '%Y-%m-%d').date()
                academic_year = get_object_or_404(AcademicYear, id=academic_year_id)

                # Prevent duplicates
                if Holiday.objects.filter(academic_year=academic_year, date=holiday_date).exists():
                    messages.warning(request, "A holiday already exists for this date.")
                else:
                    Holiday.objects.create(
                        reason=reason,
                        date=holiday_date,
                        academic_year=academic_year
                    )
                    messages.success(request, "Holiday added successfully.")
                    return redirect('academic_year_list')

            except ValueError:
                messages.error(request, "Invalid date format.")

    return render(request, 'school_management/add_holiday.html', {
        'academic_years': academic_years
    })
# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Student, Class, AcademicYear

def promote_students(request):
    classes = Class.objects.all()
    academic_years = AcademicYear.objects.all()
    selected_class_id = request.GET.get('class_id')
    students = []

    if selected_class_id:
        students = Student.objects.filter(assigned_class_id=selected_class_id)

    if request.method == 'POST':
        selected_students = request.POST.getlist('students')
        new_class_id = request.POST.get('new_class')
        new_academic_year_id = request.POST.get('new_academic_year')

        if not selected_students or not new_class_id or not new_academic_year_id:
            messages.error(request, "Please select all fields before submitting.")
            return redirect(request.path + f"?class_id={selected_class_id}")

        for student_id in selected_students:
            try:
                student = Student.objects.get(id=student_id)
                student.assigned_class_id = new_class_id
                student.academic_year_id = new_academic_year_id
                student.roll_number = None  # Triggers regeneration in `save()`
                student.save()
            except Student.DoesNotExist:
                continue

        messages.success(request, f"{len(selected_students)} students promoted successfully!")
        return redirect(request.path)

    return render(request, 'school_management/promote_students.html', {
        'classes': classes,
        'academic_years': academic_years,
        'students': students,
        'selected_class_id': int(selected_class_id) if selected_class_id else None,
    })

from django.shortcuts import render
from .models import TransferredStudent,AITimetableEntry

def transferred_student_list(request):
    students = TransferredStudent.objects.all().order_by('-transferred_on')
    return render(request, 'school_management/transferred_student_list.html', {'students': students})


def ai_timetable_setup_pg(request):
    """
    Render the form page where the user can input parameters to generate the AI timetable.
    """
    return render(request, 'school_management/Timetable/AI_Timetable_setup_pg.html')

from django.shortcuts import redirect
from django.contrib import messages





from django.shortcuts import redirect
from django.http import HttpResponseNotAllowed
from django.contrib import messages
from collections import defaultdict
from datetime import datetime, timedelta
import random

from school_app.models import (
    Class, SchoolDay, AITimetableEntry,
    Subject, Teacher_Subject_Class_Relation
)

from collections import defaultdict
from datetime import datetime, timedelta
import random

from datetime import datetime, timedelta
from collections import defaultdict
import random

from django.http import HttpResponseNotAllowed
from django.shortcuts import redirect
from django.contrib import messages

from .models import (
    AITimetableSettings,
    AITimetableEntry,
    SchoolDay,
    Class,
    Teacher_Subject_Class_Relation,
    Subject
)

def generate_timetable_with_ai(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    try:
        # Step 1: Fetch AI timetable settings from form input
        periods_per_day = int(request.POST.get('num_periods', 8))
        days_per_week = 6  # Monday to Saturday

        start_hour = int(request.POST.get('start_hour', 8))
        start_minute = int(request.POST.get('start_minute', 0))
        period_duration = int(request.POST.get('period_duration', 45))
        gap_minutes = int(request.POST.get('gap_minutes', 5))
        break_after_period_1 = int(request.POST.get('break_after_period_1', 3))
        break_after_period_2 = int(request.POST.get('break_after_period_2', 6))
        break_duration = int(request.POST.get('break_duration', 15))

        # Step 2: Store or update AI settings in the database
        AITimetableSettings.objects.all().delete()  # Ensure only one setting
        AITimetableSettings.objects.create(
            num_periods=periods_per_day,
            period_duration=period_duration,
            gap_between_periods=gap_minutes,
            break_after_period_1=break_after_period_1,
            break_after_period_2=break_after_period_2,
            break_duration=break_duration,
            start_hour=start_hour,
            start_minute=start_minute,
        )

        # Step 3: Setup required school days
        day_code_map = {
            'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed',
            'Thursday': 'Thu', 'Friday': 'Fri', 'Saturday': 'Sat'
        }
        needed_days = list(day_code_map.keys())[:days_per_week]
        needed_day_codes = [day_code_map[day] for day in needed_days]

        existing_day_codes = SchoolDay.objects.values_list('day_name', flat=True)
        for code in needed_day_codes:
            if code not in existing_day_codes:
                SchoolDay.objects.create(day_name=code)

        school_days = SchoolDay.objects.filter(day_name__in=needed_day_codes).order_by('id')
        classes = Class.objects.all()

        if not classes.exists():
            messages.error(request, "No classes found.")
            return redirect('generate_timetable_with_ai')

        if not school_days.exists():
            messages.error(request, "No school days configured.")
            return redirect('generate_timetable_with_ai')

        # Step 4: Clear previous AI timetable
        AITimetableEntry.objects.all().delete()

        # Step 5: Build class-subject-teacher map
        class_subject_teacher_map = defaultdict(dict)
        for rel in Teacher_Subject_Class_Relation.objects.all():
            for cls in rel.assigned_classes.all():
                if rel.subject and rel.teacher:
                    class_subject_teacher_map[cls.id][rel.subject.id] = rel.teacher

        # Step 6: Create or fetch break subject
        break_subject, _ = Subject.objects.get_or_create(name="Break")

        # Step 7: Generate timetable for each class
        base_time = datetime(2000, 1, 1, start_hour, start_minute)

        for cls in classes:
            class_teacher = cls.class_teacher
            class_teacher_subject_id = None

            if class_teacher and cls.id in class_subject_teacher_map:
                for subject_id, teacher in class_subject_teacher_map[cls.id].items():
                    if teacher == class_teacher:
                        class_teacher_subject_id = subject_id
                        break

            subjects = list(class_subject_teacher_map[cls.id].keys()) if cls.id in class_subject_teacher_map else []
            total_slots = days_per_week * periods_per_day
            non_first_periods_per_day = periods_per_day - 1
            total_non_first_period_slots = days_per_week * non_first_periods_per_day

            subject_distribution = []
            if subjects:
                subject_distribution = (subjects * (total_non_first_period_slots // len(subjects))) + \
                    random.sample(subjects, total_non_first_period_slots % len(subjects))
                random.shuffle(subject_distribution)

            subject_index = 0
            daily_subject_limit = defaultdict(lambda: defaultdict(int))

            for day in school_days:
                for period in range(1, periods_per_day + 1):
                    minutes_to_add = sum(
                        period_duration + gap_minutes +
                        (break_duration if p == break_after_period_1 or p == break_after_period_2 else 0)
                        for p in range(1, period)
                    )
                    current_time = base_time + timedelta(minutes=minutes_to_add)
                    start_time = current_time.time()
                    end_time = (current_time + timedelta(minutes=period_duration)).time()

                    if period == break_after_period_1 or period == break_after_period_2:
                        AITimetableEntry.objects.create(
                            class_assigned=cls,
                            subject=break_subject,
                            teacher=None,
                            school_day=day,
                            period_number=period,
                            start_time=start_time,
                            end_time=end_time
                        )
                        continue

                    if period == 1 and class_teacher_subject_id:
                        teacher = class_teacher
                        AITimetableEntry.objects.create(
                            class_assigned=cls,
                            subject_id=class_teacher_subject_id,
                            teacher=teacher,
                            school_day=day,
                            period_number=period,
                            start_time=start_time,
                            end_time=end_time
                        )
                        continue

                    if not subject_distribution:
                        continue

                    subject_id = subject_distribution[subject_index]
                    attempts = 0
                    while daily_subject_limit[cls.id][(day.id, subject_id)] >= 1 and attempts < len(subjects):
                        subject_index = (subject_index + 1) % len(subject_distribution)
                        subject_id = subject_distribution[subject_index]
                        attempts += 1

                    teacher = class_subject_teacher_map[cls.id].get(subject_id)
                    if not teacher:
                        subject_index = (subject_index + 1) % len(subject_distribution)
                        continue

                    AITimetableEntry.objects.create(
                        class_assigned=cls,
                        subject_id=subject_id,
                        teacher=teacher,
                        school_day=day,
                        period_number=period,
                        start_time=start_time,
                        end_time=end_time
                    )
                    daily_subject_limit[cls.id][(day.id, subject_id)] += 1
                    subject_index = (subject_index + 1) % len(subject_distribution)

        messages.success(request, "AI timetable generated for all classes successfully.")
        return redirect('display_ai_timetable')

    except Exception as e:
        print(f"[ERROR] Exception occurred: {e}")
        messages.error(request, f"Error generating timetable: {e}")
        return redirect('generate_timetable_with_ai')


from datetime import datetime
from django.shortcuts import render, redirect
from datetime import datetime
from django.shortcuts import render, redirect


import logging
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.shortcuts import redirect
from .models import (
    SchoolDay, AITimetableSettings, AITimetableEntry,
    TeacherAttendance, Teacher_Subject_Class_Relation,
    SubstituteAssignment, DailyTimeTableEntry
)

logger = logging.getLogger(__name__)

def generate_daily_timetable_view(request):
    if request.method != 'POST':
        return HttpResponse("Invalid request method.", status=405)

    try:
        date_str = request.POST.get('date')
        if not date_str:
            return HttpResponse("Date is required.", status=400)

        try:
            date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

        weekday = date.strftime("%a")

        try:
            school_day = SchoolDay.objects.get(day_name=weekday)
        except SchoolDay.DoesNotExist:
            return HttpResponse(f"No school day configured for {weekday}.", status=400)

        settings = AITimetableSettings.objects.last()
        if not settings:
            return HttpResponse("AI Timetable Settings not configured.", status=400)

        # Clear existing entries
        DailyTimeTableEntry.objects.filter(date=date).delete()

        timetable_entries = AITimetableEntry.objects.filter(school_day=school_day)

        period_times = []
        current_time = datetime.combine(date, datetime.min.time()).replace(
            hour=settings.start_hour, minute=settings.start_minute
        )

        for i in range(1, settings.num_periods + 1):
            start_time = current_time.time()
            end_time = (current_time + timedelta(minutes=settings.period_duration)).time()
            period_times.append((i, start_time, end_time))

            current_time += timedelta(minutes=settings.period_duration)

            if i in [settings.break_after_period_1, settings.break_after_period_2]:
                current_time += timedelta(minutes=settings.break_duration)
            else:
                current_time += timedelta(minutes=settings.gap_between_periods)

        period_time_map = {pnum: (start, end) for pnum, start, end in period_times}

        for entry in timetable_entries:
            assigned_teacher = entry.teacher
            attendance = TeacherAttendance.objects.filter(teacher=assigned_teacher, date=date).first()
            is_absent = attendance and attendance.status in ['absent', 'leave']

            if is_absent:
                substitute_relation = Teacher_Subject_Class_Relation.objects.filter(
                    subject=entry.subject,
                    assigned_classes=entry.class_assigned
                ).exclude(teacher=assigned_teacher).first()

                if substitute_relation:
                    substitute_teacher = substitute_relation.teacher
                    SubstituteAssignment.objects.create(
                        original_teacher=assigned_teacher,
                        substitute_teacher=substitute_teacher,
                        date=date,
                        subject=entry.subject,
                        classroom=entry.class_assigned,
                        period_number=entry.period_number
                    )
                    assigned_teacher = substitute_teacher
                else:
                    logger.warning(
                        f"No substitute found for class {entry.class_assigned} subject {entry.subject} on {date}"
                    )
                    assigned_teacher = None

            start_time, end_time = period_time_map.get(entry.period_number, (entry.start_time, entry.end_time))

            DailyTimeTableEntry.objects.create(
                classroom=entry.class_assigned,
                subject=entry.subject,
                teacher=assigned_teacher,
                date=date,
                period_number=entry.period_number,
                start_time=start_time,
                end_time=end_time
            )

        return redirect(f'/daily_timetable_view/?date={date}')

    except Exception as e:
        logger.error(f"Error generating daily timetable: {e}", exc_info=True)
        return HttpResponse("Internal server error.", status=500)


from .models import (
    AITimetableEntry, AITimetableSettings, Class, TeacherAttendance, Teacher
)

from collections import defaultdict
from django.shortcuts import render
from .models import Class, AITimetableEntry, SchoolDay

def display_ai_timetable(request):
    selected_date = request.GET.get('date', date.today().isoformat())
    today = date.today()
    # Get all classes for dropdown
    all_classes = Class.objects.all()
    
    class_id = request.GET.get('class_id')
    class_obj = Class.objects.filter(id=class_id).first() if class_id else Class.objects.first()

    if not class_obj:
        return render(request, 'school_management/Timetable/AI_timetable_display.html', {
            'error': 'No class found to display timetable.',
            'all_classes': all_classes,
        })

    # Fetch timetable entries for selected class
    entries = AITimetableEntry.objects.filter(class_assigned=class_obj)\
        .select_related('subject', 'teacher', 'school_day')\
        .order_by('school_day__id', 'period_number')

    # All days in order
    school_days = list(SchoolDay.objects.all().order_by('id'))

    # Find max period number
    max_period = 0
    for e in entries:
        if e.period_number > max_period:
            max_period = e.period_number
    max_periods = list(range(1, max_period + 1))

    # Create period timings list of dicts (period number and formatted timing string)
    period_timings = []
    for period in max_periods:
        sample_entry = next((e for e in entries if e.period_number == period), None)
        if sample_entry:
            timing_str = f"{sample_entry.start_time.strftime('%H:%M')} - {sample_entry.end_time.strftime('%H:%M')}"
        else:
            timing_str = "--:--"
        period_timings.append({
            'period': period,
            'timing': timing_str,
        })

    # Build timetable lookup map (day_name, period_number) -> entry
    timetable_map = {
        (e.school_day.day_name, e.period_number): e for e in entries
    }

    # Build rows for each day
    timetable_rows = []
    for day in school_days:
        row_entries = [
            timetable_map.get((day.day_name, p), None)
            for p in max_periods
        ]
        timetable_rows.append({
            'day': day,
            'entries': row_entries,
        })

    # Subject-wise period count (excluding "Break")
    subject_hours = defaultdict(int)
    for entry in entries:
        if entry.subject and entry.subject.name.lower() != 'break':
            subject_hours[entry.subject.name] += 1

    # Create list for template
    subject_summary = [
        {'subject': subject, 'periods': count}
        for subject, count in subject_hours.items()
    ]

    context = {
        'all_classes': all_classes,
        'class_obj': class_obj,
        'period_timings': period_timings,
        'timetable_rows': timetable_rows,
        'subject_summary': subject_summary,
          "selected_date": selected_date,
        "today"  : today
    }

    return render(request, 'school_management/Timetable/AI_timetable_display.html', context)


from datetime import datetime, timedelta
from django.shortcuts import render, redirect
import logging

from .models import (
    AITimetableEntry, DailyTimeTableEntry, TeacherAttendance,
    SubstituteAssignment, SchoolDay, Teacher_Subject_Class_Relation,
    AITimetableSettings
)

logger = logging.getLogger(__name__)

import logging
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.shortcuts import redirect
from django.views.decorators.http import require_POST
from .models import (
    AITimetableSettings,
    SchoolDay,
    AITimetableEntry,
    DailyTimeTableEntry,
    TeacherAttendance,
    Teacher_Subject_Class_Relation,
    SubstituteAssignment,
)

logger = logging.getLogger(__name__)

@require_POST
def generate_daily_timetable_view(request):
    date_str = request.POST.get('date')
    if not date_str:
        return HttpResponse("Date is required.", status=400)

    try:
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    weekday = date.strftime("%a")

    try:
        school_day = SchoolDay.objects.get(day_name=weekday)
    except SchoolDay.DoesNotExist:
        return HttpResponse(f"No school day configured for {weekday}.", status=404)

    settings = AITimetableSettings.objects.last()
    if not settings:
        return HttpResponse("AI Timetable Settings not configured.", status=500)

    # Clear existing entries
    DailyTimeTableEntry.objects.filter(date=date).delete()

    timetable_entries = AITimetableEntry.objects.filter(school_day=school_day)

    # Calculate start and end times for each period
    period_times = []
    current_time = datetime.combine(date, datetime.min.time()).replace(
        hour=settings.start_hour, minute=settings.start_minute
    )

    for i in range(1, settings.num_periods + 1):
        start_time = current_time.time()
        end_time = (current_time + timedelta(minutes=settings.period_duration)).time()
        period_times.append((i, start_time, end_time))

        current_time += timedelta(minutes=settings.period_duration)

        if i in [settings.break_after_period_1, settings.break_after_period_2]:
            current_time += timedelta(minutes=settings.break_duration)
        else:
            current_time += timedelta(minutes=settings.gap_between_periods)

    period_time_map = {pnum: (start, end) for pnum, start, end in period_times}

    for entry in timetable_entries:
        assigned_teacher = entry.teacher
        attendance = TeacherAttendance.objects.filter(teacher=assigned_teacher, date=date).first()
        is_absent = attendance and attendance.status in ['absent', 'leave']

        if is_absent:
            substitute_relation = Teacher_Subject_Class_Relation.objects.filter(
                subject=entry.subject,
                assigned_classes=entry.class_assigned
            ).exclude(teacher=assigned_teacher).first()

            if substitute_relation:
                substitute_teacher = substitute_relation.teacher
                SubstituteAssignment.objects.create(
                    original_teacher=assigned_teacher,
                    substitute_teacher=substitute_teacher,
                    date=date,
                    subject=entry.subject,
                    classroom=entry.class_assigned,
                    period_number=entry.period_number
                )
                assigned_teacher = substitute_teacher
            else:
                logger.warning(f"No substitute found for class {entry.class_assigned} subject {entry.subject} on {date}")
                assigned_teacher = None

        start_time, end_time = period_time_map.get(entry.period_number, (entry.start_time, entry.end_time))

        DailyTimeTableEntry.objects.create(
            classroom=entry.class_assigned,
            subject=entry.subject,
            teacher=assigned_teacher,
            date=date,
            period_number=entry.period_number,
            start_time=start_time,
            end_time=end_time
        )

    return redirect(f'/daily_timetable_view/?date={date}')

from collections import defaultdict
from datetime import date
from django.shortcuts import render
from .models import (
    AITimetableEntry, TeacherAttendance, AITimetableSettings,
    Class, Teacher, SubstituteAssignment
)
from django.http import HttpResponse
from django.shortcuts import render
from datetime import date
from collections import defaultdict
from .models import (
    AITimetableEntry, TeacherAttendance, AITimetableSettings,
    SubstituteAssignment, Class, Teacher
)

def daily_timetable_view(request):
    selected_date_str = request.GET.get('date')
    
    try:
        selected_date = date.fromisoformat(selected_date_str) if selected_date_str else date.today()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.", status=400)

    weekday_map = {
        0: 'Mon', 1: 'Tue', 2: 'Wed', 3: 'Thu',
        4: 'Fri', 5: 'Sat', 6: 'Sun'
    }
    weekday_short = weekday_map[selected_date.weekday()]

    # Get timetable entries for the day
    entries = AITimetableEntry.objects.filter(
        school_day__day_name=weekday_short
    ).select_related('class_assigned', 'subject', 'teacher')

    absent_or_leave_teacher_ids = set(
        TeacherAttendance.objects.filter(date=selected_date, status__in=['absent', 'leave'])
        .values_list('teacher_id', flat=True)
    )

    try:
        settings = AITimetableSettings.objects.latest('id')
        break_periods = list(filter(None, [settings.break_after_period_1, settings.break_after_period_2]))
        num_periods = settings.num_periods
    except AITimetableSettings.DoesNotExist:
        break_periods = [3, 6]
        num_periods = 8

    period_range = range(1, num_periods + 1)

    timetable_matrix = defaultdict(dict)

    for entry in entries:
        class_name = entry.class_assigned.name
        period = entry.period_number
        subject_name = entry.subject.name if entry.subject else "N/A"
        teacher = entry.teacher

        if teacher is None:
            teacher_display = "Not Assigned"
        elif teacher.id in absent_or_leave_teacher_ids:
            substitute = SubstituteAssignment.objects.filter(
                original_teacher=teacher,
                date=selected_date,
                period_number=period,
                subject=entry.subject,
                classroom=entry.class_assigned
            ).select_related('substitute_teacher').first()

            if substitute and substitute.substitute_teacher:
                teacher_display = f"{teacher.name} (Absent → {substitute.substitute_teacher.name})"
            else:
                teacher_display = f"{teacher.name} (Absent/Leave)"
        else:
            teacher_display = teacher.name

        timetable_matrix[class_name][period] = f"{subject_name} / {teacher_display}"

    all_class_objs = Class.objects.all()
    for cls in all_class_objs:
        if cls.name not in timetable_matrix:
            timetable_matrix[cls.name] = {}

    all_classes = sorted(timetable_matrix.keys())

    table_rows = []
    for class_name in all_classes:
        row = [class_name]
        for period in period_range:
            if period in break_periods:
                row.append("Break")
            else:
                row.append(timetable_matrix[class_name].get(period, "No Class"))
        table_rows.append(row)

    period_headers = [f"P{p}" for p in period_range]

    all_teachers = Teacher.objects.all()
    present_teachers = set(all_teachers.exclude(id__in=absent_or_leave_teacher_ids))

    busy_teachers_by_period = defaultdict(set)
    for entry in entries:
        if entry.teacher and entry.teacher.id not in absent_or_leave_teacher_ids:
            busy_teachers_by_period[entry.period_number].add(entry.teacher)

    free_teachers_table = []
    for period in period_range:
        if period in break_periods:
            free_teachers_table.append(("Break", []))
        else:
            busy_teachers = busy_teachers_by_period.get(period, set())
            free_teachers = sorted(present_teachers - busy_teachers, key=lambda t: t.name)
            free_teachers_table.append((f"P{period}", [t.name for t in free_teachers]))

    context = {
        "selected_date": selected_date,
        "period_headers": period_headers,
        "table_rows": table_rows,
        "free_teachers_table": free_teachers_table,
        "today": date.today(),
    }

    return render(request, 'school_management/Timetable/daily_timetable.html', context)

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Class, Teacher, Teacher_Subject_Class_Relation

def assign_class_teachers_view(request):
    classes = Class.objects.select_related('class_teacher').all()

    # Build class-wise teacher lists
    relations = Teacher_Subject_Class_Relation.objects.select_related('teacher').prefetch_related('assigned_classes')

    class_teachers_dict = {}
    for relation in relations:
        if relation.teacher:
            for cls in relation.assigned_classes.all():
                class_teachers_dict.setdefault(cls.id, set()).add(relation.teacher)

    class_teacher_data = []
    for cls in classes:
        teachers_for_class = sorted(
            class_teachers_dict.get(cls.id, []),
            key=lambda t: t.name
        )
        class_teacher_data.append({
            'class': cls,
            'teachers': teachers_for_class
        })

    if request.method == 'POST':
        changes_made = False

        for item in class_teacher_data:
            cls = item['class']
            teacher_id = request.POST.get(f'class_teacher_{cls.id}')
            if teacher_id:
                try:
                    teacher_id = int(teacher_id)
                    if cls.class_teacher_id != teacher_id:
                        cls.class_teacher_id = teacher_id
                        cls.save()
                        changes_made = True
                except ValueError:
                    continue

        if changes_made:
            messages.success(request, "Class teachers assigned successfully.")
        else:
            messages.info(request, "No changes made.")

        return redirect('assign_class_teachers')

    return render(request, 'school_management/principal/class_teacher_assignment.html', {
        'class_teacher_data': class_teacher_data
    })
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Class, Teacher

def assign_class_teachers(request):
    # Get all classes and all teachers (or filter teachers if needed)
    classes = Class.objects.all()
    teachers = Teacher.objects.all()

    # Prepare the data structure for template
    class_teacher_data = []
    for c in classes:
        class_teacher_data.append({
            'class': c,
            'teachers': teachers,
        })

    if request.method == 'POST':
        # Iterate over classes to update assigned teachers
        for c in classes:
            teacher_id = request.POST.get(f'class_teacher_{c.id}')
            if teacher_id:
                try:
                    teacher = Teacher.objects.get(id=teacher_id)
                except Teacher.DoesNotExist:
                    teacher = None
            else:
                teacher = None
            c.class_teacher = teacher
            c.save()
        messages.success(request, "Class teachers assigned successfully.")
        return redirect('assign_class_teachers')

    return render(request, 'school_management/principal/class_teacher_assignment.html', {
        'class_teacher_data': class_teacher_data
    })


# views.py

from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date
from .models import Teacher, TeacherAttendance
from datetime import date
from datetime import date
from django.shortcuts import render, redirect
from .models import Teacher, TeacherAttendance

def mark_teacher_attendance_view(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date') or date.today().isoformat()

        teachers = Teacher.objects.all()
        for teacher in teachers:
            status = request.POST.get(f'status_{teacher.id}', 'present')
            reason = request.POST.get(f'reason_{teacher.id}', '').strip()

            attendance, created = TeacherAttendance.objects.get_or_create(
                teacher=teacher,
                date=selected_date,
                defaults={'status': status, 'reason': reason if status in ['absent', 'leave'] else ''}
            )
            if not created:
                attendance.status = status
                attendance.reason = reason if status in ['absent', 'leave'] else ''
                attendance.save()

        return redirect('mark_teacher_attendance')

    else:
        teachers = Teacher.objects.all()
        today = date.today().isoformat()
        # Get attendance for today or selected date (default to today)
        today_attendance_qs = TeacherAttendance.objects.filter(date=today).select_related('teacher')
        attendance_map = {att.teacher_id: att for att in today_attendance_qs}

        # Prepare teacher data with attendance info for template
        teacher_data = []
        for teacher in teachers:
            att = attendance_map.get(teacher.id)
            teacher_data.append({
                'teacher': teacher,
                'status': att.status if att else 'present',
                'reason': att.reason if att else '',
            })

        return render(request, 'school_management/Timetable/mark_teacher_attendance_manual.html', {
            'teacher_data': teacher_data,
            'selected_date': today,
        })

from django.shortcuts import render, redirect
from django import forms
from .models import SchoolProfile
from django.contrib import messages

from django.shortcuts import render
from django.http import HttpResponse
from .models import SchoolProfile

def school_profile_view(request):
    profile = SchoolProfile.objects.first()
    if not profile:
        profile = SchoolProfile.objects.create()  # Create default if none exists

    if request.method == 'POST':
        profile.name = request.POST.get('name', profile.name)
        profile.address = request.POST.get('address', profile.address)
        profile.phone_number = request.POST.get('phone_number', profile.phone_number)
        profile.email = request.POST.get('email', profile.email)
        profile.principal_name = request.POST.get('principal_name', profile.principal_name)

        if request.FILES.get('logo'):
            profile.logo = request.FILES['logo']

        profile.save()
        return HttpResponse("School profile updated successfully.")

    return render(request, 'school_management/admin/school_profile.html', {'profile': profile})
